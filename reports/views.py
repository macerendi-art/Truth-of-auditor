"""Laporan siap setor: ekspor XLSX level batch (semua relasi sekaligus).

Beda dengan export_run (bukti kerja satu relasi, semua bucket), ekspor batch
menjawab pertanyaan atasan: "hari itu selisihnya berapa dan baris mana saja
yang masih terbuka?" — ringkasan DP/WD + hanya baris tidak_cocok/perlu_tinjau.
"""
import io

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font

from reconciliation.models import MatchResult, ReconBatch
from web.access import tokos_for
from web.templatetags.web_extras import reason_label

_BARIS_HEAD = [
    "Relasi", "Jenis", "Ticket", "User", "Nama", "Player Bank",
    "Jumlah", "Waktu", "Alasan", "Detail",
]


def _isi_baris(ws, results):
    ws.append(_BARIS_HEAD)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in results:
        left = r.left
        ws.append([
            r.run.get_relation_display(),
            (left.jenis or "").upper() if left else "",
            left.ticket_no if left else "",
            left.username if left else "",
            left.counterparty if left else "",
            (left.raw or {}).get("Player Bank", "") if left else "",
            float(left.amount) if left else "",
            left.occurred_at.strftime("%d/%m/%Y %H:%M") if left and left.occurred_at else "",
            reason_label(r.reason_code),
            r.reason_detail,
        ])


@login_required
def export_batch(request, pk):
    batch = get_object_or_404(ReconBatch, pk=pk, toko__in=tokos_for(request.user))
    s = batch.summary or {}
    dp, wd = s.get("dp") or {}, s.get("wd") or {}
    bk = s.get("buckets") or {}

    wb = Workbook()
    ring = wb.active
    ring.title = "Ringkasan"
    label_window = "semua tanggal"
    if batch.date_from and batch.date_to:
        label_window = (
            batch.date_from.strftime("%d/%m/%Y")
            if batch.date_from == batch.date_to
            else f"{batch.date_from.strftime('%d/%m/%Y')} – {batch.date_to.strftime('%d/%m/%Y')}"
        )
    for label, val in [
        ("Toko", batch.toko.name if batch.toko else ""),
        ("Tanggal data", label_window),
        ("Dibuat", batch.created_at.strftime("%d/%m/%Y %H:%M")),
        ("Toleransi", f"{batch.tolerance.name} (±{batch.tolerance.date_window_days} hari)"),
        ("", ""),
        ("DP panel", dp.get("panel", 0)),
        ("DP uang matched", dp.get("money_matched", dp.get("money", 0))),
        ("DP selisih", dp.get("selisih", 0)),
        ("WD panel", wd.get("panel", 0)),
        ("WD uang matched", wd.get("money_matched", wd.get("money", 0))),
        ("WD selisih", wd.get("selisih", 0)),
        ("", ""),
        ("Cocok", bk.get("cocok", 0)),
        ("Perlu tinjau", bk.get("perlu_tinjau", 0)),
        ("Tidak cocok", bk.get("tidak_cocok", 0)),
    ]:
        ring.append([label, val])
    for row in ring["A"]:
        row.font = Font(bold=True)

    base = (
        MatchResult.objects.filter(run__batch=batch)
        .select_related("left", "run")
        .order_by("run__relation", "left__occurred_at")
    )
    _isi_baris(wb.create_sheet("Selisih"), base.filter(bucket=MatchResult.Bucket.TIDAK))
    _isi_baris(wb.create_sheet("Perlu Tinjau"), base.filter(bucket=MatchResult.Bucket.TINJAU))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama = f"laporan_batch{batch.pk}_{(batch.date_to or batch.created_at.date()).isoformat()}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{nama}"'
    return resp

"""Ekspor XLSX level batch: laporan selisih siap setor per hari data.

Auditor butuh satu file untuk atasan: ringkasan DP/WD + daftar baris yang
masih selisih / perlu tinjau, lintas semua relasi batch itu.
"""
import io
from datetime import date, datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from reconciliation.models import MatchResult, MatchRun, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction

User = get_user_model()


class ExportBatchTests(TestCase):
    def setUp(self):
        self.lbs = Toko.objects.get(key="lbs")
        self.tol = ToleranceProfile.objects.get(name="Default")
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.adm = User.objects.create_user("adm", password="pw123456", role="admin")
        self.client.login(username="adm", password="pw123456")
        self.client.post(reverse("set_toko"), {"toko_id": self.lbs.id})
        self.up = Upload.objects.create(source_type=self.panel, toko=self.lbs)
        self.batch = ReconBatch.objects.create(
            toko=self.lbs, tolerance=self.tol,
            date_from=date(2026, 6, 28), date_to=date(2026, 6, 28),
            summary={"dp": {"panel": 100, "selisih": 50}, "wd": {"panel": 0, "selisih": 0},
                     "buckets": {"cocok": 1, "perlu_tinjau": 1, "tidak_cocok": 1}},
        )
        self.run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=self.batch,
            date_from=date(2026, 6, 28), date_to=date(2026, 6, 28),
            summary={"cocok": 1, "perlu_tinjau": 1, "tidak_cocok": 1},
        )
        self._n = 0
        self._res("selisihuser", MatchResult.Bucket.TIDAK, "no_money")
        self._res("tinjauuser", MatchResult.Bucket.TINJAU, "weak_name")
        self._res("cocokuser", MatchResult.Bucket.COCOK, "")

    def _res(self, username, bucket, reason):
        self._n += 1
        tx = Transaction.objects.create(
            upload=self.up, source_type=self.panel, toko=self.lbs, jenis="dp",
            amount=Decimal("50000"), money_delta=Decimal("50000"), username=username,
            occurred_at=datetime(2026, 6, 28, 10, 0), row_hash=f"xb{self._n}",
        )
        return MatchResult.objects.create(
            run=self.run, bucket=bucket, left=tx, score=0, reason_code=reason,
        )

    def _load(self):
        r = self.client.get(reverse("export_batch", args=[self.batch.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        return load_workbook(io.BytesIO(r.content))

    def test_sheet_ringkasan_dan_baris_terbuka(self):
        wb = self._load()
        self.assertIn("Ringkasan", wb.sheetnames)
        self.assertIn("Selisih", wb.sheetnames)
        self.assertIn("Perlu Tinjau", wb.sheetnames)
        ring = [str(c.value) for row in wb["Ringkasan"].iter_rows() for c in row if c.value]
        self.assertTrue(any("LBS" in v for v in ring))
        selisih = [c.value for row in wb["Selisih"].iter_rows() for c in row]
        self.assertIn("selisihuser", selisih)
        self.assertNotIn("cocokuser", selisih)
        tinjau = [c.value for row in wb["Perlu Tinjau"].iter_rows() for c in row]
        self.assertIn("tinjauuser", tinjau)

    def test_rbac_toko_lain_404(self):
        aud = User.objects.create_user("aud", password="pw123456", role="auditor")
        lain = Toko.objects.exclude(pk=self.lbs.pk).first()
        aud.allowed_tokos.add(lain)
        self.client.login(username="aud", password="pw123456")
        r = self.client.get(reverse("export_batch", args=[self.batch.pk]))
        self.assertEqual(r.status_code, 404)

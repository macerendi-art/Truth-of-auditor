"""Parser bank (sumber UANG). Format beda tiap bank.

- BRI : CSV, header baris 1, kolom MUTASI_DEBET/KREDIT + SALDO_AKHIR_MUTASI.
- BCA : CSV, ada preamble; header 'Tanggal,Keterangan,Cabang,Jumlah,,Saldo' (DB/CR).
        Ekspor MyBCA kadang ; / tab / UTF-16 — auto-deteksi.
- Mandiri: xlsx e-Statement; header 2 baris; tiap transaksi 2 baris (tgl lalu jam);
  angka format ID (1.000,00).
"""
import csv
import io
import re
from decimal import Decimal

import openpyxl

from .base import (
    BaseParser,
    parse_decimal,
    parse_dt,
    read_csv_raw,
    row_hash,
    rows_to_dicts,
)
from .fee_rules import is_admin_fee


def _jenis_from_money(money):
    return "depo" if money > 0 else "wd" if money < 0 else "lainnya"


# Baris biaya transaksi BCA ("BI-FAST DB BIAYA TXN ... 2.500"): tiap WD nyata
# berpasangan dengan satu baris fee ini. Ditandai 'admin' agar dikecualikan dari
# uang WD & pencocokan (baris tetap disimpan untuk audit).
BCA_FEE_RE = re.compile(r"BIAYA\s+TXN", re.IGNORECASE)


def is_bca_fee(desc):
    return bool(BCA_FEE_RE.search(str(desc or "")))


# Fee transfer BRIVA (WD e-wallet via BRI): tiap transfer berpasangan baris
# debit Rp1.000 ber-SEQ & deskripsi IDENTIK. Bukti data WLG 01-10 Jul 2026:
# 182 pasangan persis, nol fee yatim, nol nominal BRIVA lain <10rb.
# Wajib pola 'BRIVA<digit>' (bukan substring) agar penerima transfer biasa
# yang kebetulan bernama mengandung 'BRIVA' tidak ikut tertandai.
BRIVA_FEE = Decimal("1000")
BRIVA_DESC_RE = re.compile(r"BRIVA\s*\d{3}", re.IGNORECASE)


def is_briva_fee(desc, money):
    return money == -BRIVA_FEE and bool(BRIVA_DESC_RE.search(str(desc or "")))


# ---------------------------------------------------------------------------
# Isolasi nama (Task 4). Urutan wajib: buang teks struktural per-sumber DULU,
# baru nama dinormalisasi (clean_name) di engine saat fuzzy matching.
# Angka/simbol masih dibutuhkan di tahap ini untuk mengenali pola struktural.
# ---------------------------------------------------------------------------

# --- BCA (dipakai CSV & PDF) ---
# Nama via baris lanjutan e-wallet: 'TRFDN-<nama>ESPAY DEBIT INDONE' (bisa menempel).
BCA_TRFDN_RE = re.compile(r"TRFDN-\s*(.+?)\s*(?:ESPAY\s+DEBIT\s+INDONE\S*|ESPAY|$)")
# Kode transaksi: '2706/FTSCY/WS95271', '2606/FTFVA/WS9501139010/DANA', dst.
BCA_CODE_RE = re.compile(r"\b\d{3,4}/[A-Z]+/\S+")
# Nominal menempel ke nama: '100000.00M. YULIANSAR SIREG' -> nama di belakang nominal.
BCA_GLUED_AMT_RE = re.compile(r"\d[\d,.]*\.\d{2}\s*(.*)$")
# Label/kata struktural yang bukan bagian nama (frasa dulu, baru kata tunggal).
BCA_NOISE_RE = re.compile(
    r"TRSF E-BANKING|BI-?FAST|SWITCHING|ESPAY\s+DEBIT\s+INDONE\S*|DEBIT\s+INDONE\S*|ESPAY"
    r"|Web BRILink|MyBCA|\bKBI\b|\bTOPUP\b|\bTANGGAL\b|\bTRANSFER\b|\bBIAYA\b|\bTXN\b"
    r"|\bTRF\b|\bDR\b|\bKE\b|\bCR\b|\bDB\b"
)


def extract_bca_name(text):
    """Isolasi nama orang dari keterangan BCA: buang kode transaksi, label
    struktural, nominal menempel, & nomor rekening/HP. Baris tanpa nama
    (mis. topup DANA/GOPAY hanya nomor HP) -> '' (jangan dikarang)."""
    s = re.sub(r"\s+", " ", str(text or "")).strip()
    trfdn = BCA_TRFDN_RE.search(s)
    s = BCA_TRFDN_RE.sub(" ", s)
    s = BCA_CODE_RE.sub(" ", s)
    s = BCA_NOISE_RE.sub(" ", s)
    m = BCA_GLUED_AMT_RE.search(s)
    if m:  # nama menempel di belakang nominal -> ambil bagian setelahnya
        s = m.group(1)
    # Sisa token murni angka/simbol = nomor rekening/HP/kode -> bukan nama.
    toks = [t for t in s.split() if re.search(r"[A-Za-z]", t) and not re.search(r"\d", t)]
    name = " ".join(toks).strip(" -.,:/")
    if not name and trfdn:  # fallback: nama dari baris lanjutan TRFDN
        name = trfdn.group(1).strip(" -.,:/")
    return re.sub(r"\s+", " ", name).strip()


# --- Mandiri ---
# Prefiks: 'Transfer dari/ke ...', 'Transfer BI Fast Dari/Ke ...',
# 'Transfer antar Mandiri DARI ...' (+ opsional 'Bank lain').
MANDIRI_PREFIX_RE = re.compile(
    r"^Transfer\s+(?:BI\s*Fast\s+)?(?:dari|ke|antar)\s+(?:Mandiri\s+(?:dari|ke)\s+)?(?:Bank\s+lain\s+)?",
    re.IGNORECASE,
)
# Nama bank pengirim/penerima (terpanjang dulu agar 'BANK MANDIRI TASPEN'
# tidak terpotong jadi 'BANK MANDIRI' + sisa 'TASPEN').
MANDIRI_BANK_NAMES = (
    "BANK MANDIRI TASPEN", "SUPER BANK INDONESIA", "SEABANK INDONESIA",
    "BANK RAKYAT INDONESIA", "BANK CENTRAL ASIA", "BANK NEGARA INDONESIA",
    "BANK SYARIAH INDONESIA", "BANK NEO COMMERCE", "BANK CIMB NIAGA",
    "BANK MANDIRI", "BANK DANAMON", "BANK PERMATA", "BANK JAGO", "BANK MEGA",
    "BANK BTPN", "BANK BNI", "BANK BRI", "BANK BCA", "BANK BTN", "BANK BJB",
    "BCA DIGITAL", "CIMB NIAGA", "SUPERBANK", "ALLO BANK", "SEABANK",
    "BCA", "BRI", "BNI", "BTN", "BSI",
)
# Ekor struktural GoPay/fee: buang sampai akhir teks.
MANDIRI_TAIL_RES = (
    re.compile(r"\bGoPay\s+Bank\s+Transfer\b.*$", re.IGNORECASE),
    re.compile(r"\bTransfer\s+Fee\b.*$", re.IGNORECASE),
)


def extract_mandiri_name(text):
    """Isolasi nama dari Keterangan Mandiri e-statement: buang prefiks
    'Transfer dari/ke <BANK>', nama bank, dan nomor rekening/referensi di ekor.
    Baris biaya/pembayaran tanpa nama -> ''. """
    s = re.sub(r"\s+", " ", str(text or "")).strip()
    if not s.lower().startswith("transfer"):
        return ""  # 'Biaya ...' / 'Pembayaran GoPay Customer <HP>' -> tanpa nama
    stripped = MANDIRI_PREFIX_RE.sub("", s)
    if stripped == s:  # bukan pola transfer yang dikenal -> jangan menebak
        return ""
    s = stripped.strip()
    upper = s.upper()
    for bank in MANDIRI_BANK_NAMES:  # buang nama bank di depan (batas kata)
        if upper == bank or upper.startswith(bank + " "):
            s = s[len(bank):].strip()
            break
    s = re.sub(r"^DANA-\s*", "", s)  # e-wallet: 'DANA-<nama>' menempel
    for tail_re in MANDIRI_TAIL_RES:
        s = tail_re.sub("", s).strip()
    # Ekor nomor rekening/HP/referensi (mengandung angka) + tanda '-'.
    toks = s.split()
    while toks and (re.search(r"\d", toks[-1]) or toks[-1] in ("-", "transfer")):
        toks.pop()
    return " ".join(toks).strip(" -.,:/")


# --- BRI ---
# 'NBMB <pengirim> TO <penerima> ESB:NBMB:...' (bentuk lazim) ATAU, ditemukan pada
# mutasi WLG 25-07, 'NBMB <pengirim> TO <penerima>' TANPA trailer ESB sama sekali.
# Grup nama tetap lazy & berhenti di kemunculan ' ESB' PERTAMA (mempertahankan
# perilaku lama persis) — bila tak ada ' ESB', berhenti di akhir string.
# Modul konstan (satu sumber): dipakai parser di sini DAN fallback tampilan
# query-time di web/views.py (baris lama yang sudah tersimpan tanpa counterparty).
# BRIVA ('...NBMBAxxxx...' nempel tanpa spasi) & fee BI-Fast ('NBMB:X' titik dua
# nempel) sengaja TIDAK cocok — keduanya tak punya pola literal 'NBMB ' + ' TO '.
NBMB_RE = re.compile(r"NBMB (.+?) TO (.+?)(?: ESB|$)")


class BRIParser(BaseParser):
    source_key = "bank"

    def parse(self, path, flow=""):
        rows = read_csv_raw(path)
        _, dicts = rows_to_dicts(rows, 0)
        out = []
        for r in dicts:
            debit = parse_decimal(r.get("MUTASI_DEBET"))
            credit = parse_decimal(r.get("MUTASI_KREDIT"))
            money = credit - debit
            occurred = parse_dt(r.get("TGL_TRAN"))
            desc = str(r.get("DESK_TRAN", "") or "")
            m = NBMB_RE.search(desc)
            sender, receiver = (m.group(1).strip(), m.group(2).strip()) if m else ("", "")
            counterparty = sender if money > 0 else receiver
            seq = str(r.get("SEQ", "") or "").strip()
            row = {
                "source_type": "bank",
                "occurred_at": occurred,
                "posted_date": occurred.date() if occurred else None,
                "jenis": "admin"
                if (money < 0 and (is_briva_fee(desc, money)
                                   or is_admin_fee("bri", desc, abs(money))))
                else _jenis_from_money(money),
                "amount": abs(money),
                "credit_delta": Decimal("0"),
                "money_delta": money,
                "fee": Decimal("0"),
                "bonus": Decimal("0"),
                "balance_after": parse_decimal(r.get("SALDO_AKHIR_MUTASI")),
                "ticket_no": "",
                "username": "",
                "reference": seq,
                "counterparty": counterparty,
                "description": desc,
                "raw": {k: ("" if v is None else str(v)) for k, v in r.items()},
            }
            row["row_hash"] = row_hash("bri", [r.get("NOREK", ""), seq, occurred, money])
            out.append(row)
        return out


class BCACSVParser(BaseParser):
    source_key = "bank"

    # Header kolom — ekspor klasik + varian MyBCA / EN
    _HDR_TGL = ("tanggal", "date", "tgl", "transaction date")
    _HDR_SALDO = ("saldo", "balance", "saldo (idr)", "ending balance")
    _HDR_JUMLAH = ("jumlah", "amount", "mutasi", "nominal", "transaction amount")
    _HDR_KET = ("keterangan", "description", "deskripsi", "narration", "uraian")

    def parse(self, path, flow=""):
        rows, delim = self._baca_csv_otomatis(path)
        hidx = None
        for i, r in enumerate(rows):
            cells = [str(c).strip() for c in r]
            # Preamble BCA: baris 'Nama,=,NIJUN' sebelum header -> pemilik rekening.
            if not self.meta.get("owner_name") and cells and self._sel_nama(cells):
                owner = self._ambil_owner(cells)
                if owner:
                    self.meta["owner_name"] = owner
            if self._is_header(cells):
                hidx = i
                break
        if hidx is None:
            # Header tak ketemu — jangan diam (KIGAR SHU MING: 0 baris + owner dari nama file)
            if any(any(str(c).strip() for c in r) for r in rows):
                raise ValueError(
                    "Mutasi BCA CSV tidak dikenali: header Tanggal/Saldo tidak ditemukan "
                    f"(delimiter coba={delim!r}). Kirim sample ke pengembang."
                )
            return []
        _, dicts = rows_to_dicts(rows, hidx)
        # Normalisasi alias kolom → kunci klasik
        dicts = [self._norm_row(d) for d in dicts]
        out = []
        for r in dicts:
            jumlah = parse_decimal(r.get("Jumlah"))
            dbcr = ""
            for v in r.values():
                vv = str(v).strip().upper()
                if vv in ("DB", "CR", "D", "C", "DEBIT", "CREDIT"):
                    dbcr = "DB" if vv in ("DB", "D", "DEBIT") else "CR"
                    break
            # Beberapa ekspor: jumlah negatif = DB, positif = CR, tanpa kolom DB/CR
            if not dbcr:
                raw_j = str(r.get("Jumlah") or "").strip()
                if raw_j.startswith("-") or (jumlah is not None and jumlah < 0):
                    dbcr = "DB"
                    jumlah = abs(jumlah or Decimal("0"))
                else:
                    dbcr = "CR"
            money = jumlah if dbcr == "CR" else -abs(jumlah or Decimal("0"))
            occurred = parse_dt(r.get("Tanggal"), dayfirst=True)
            if occurred is None:  # skip baris ringkasan (Saldo Awal/Akhir/Mutasi)
                continue
            desc = str(r.get("Keterangan", "") or "")
            row = {
                "source_type": "bank",
                "occurred_at": occurred,
                "posted_date": occurred.date() if occurred else None,
                "jenis": "admin" if is_bca_fee(desc) else _jenis_from_money(money),
                "amount": abs(money),
                "credit_delta": Decimal("0"),
                "money_delta": money,
                "fee": Decimal("0"),
                "bonus": Decimal("0"),
                "balance_after": parse_decimal(r.get("Saldo")),
                "ticket_no": "",
                "username": "",
                "reference": "",
                "counterparty": extract_bca_name(desc),
                "description": desc,
                "raw": {k: ("" if v is None else str(v)) for k, v in r.items()},
            }
            row["row_hash"] = row_hash("bca", [occurred, money, r.get("Saldo", ""), desc[:60]])
            out.append(row)
        if not out and dicts:
            raise ValueError(
                "Mutasi BCA CSV: header ketemu tapi 0 baris bertanggal "
                f"({len(dicts)} baris data). Cek kolom Tanggal/Jumlah — kirim sample."
            )
        return out

    def _baca_csv_otomatis(self, path):
        """Baca CSV BCA dengan encoding + delimiter otomatis.

        MyBCA / Excel ID sering `;` atau tab; kadang UTF-16. Owner dari nama file
        dulu menutupi kegagalan header (KIGAR SHU MING rows_parsed=0).
        """
        with open(path, "rb") as f:
            raw = f.read()
        if not raw.strip():
            return [], ","

        text = None
        used_enc = "utf-8-sig"
        for enc in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"):
            try:
                cand = raw.decode(enc)
            except UnicodeDecodeError:
                continue
            # UTF-16 mis-decode as latin often has many NULs
            if enc.startswith("utf-16") or "\x00" not in cand[:200]:
                text = cand
                used_enc = enc
                if enc.startswith("utf-16") or self._skor_header_text(cand) > 0:
                    break
        if text is None:
            text = raw.decode("utf-8-sig", errors="replace")

        best_rows, best_delim, best_score = None, ",", -1
        for delim in (",", ";", "\t", "|"):
            try:
                rows = list(csv.reader(io.StringIO(text), delimiter=delim))
            except csv.Error:
                continue
            score = self._skor_rows(rows)
            if score > best_score:
                best_score, best_rows, best_delim = score, rows, delim
        if best_rows is None:
            best_rows = read_csv_raw(path)
        return best_rows, best_delim

    def _skor_header_text(self, text: str) -> int:
        t = text.casefold()
        s = 0
        if "tanggal" in t or "date" in t:
            s += 1
        if "saldo" in t or "balance" in t:
            s += 1
        if "keterangan" in t or "description" in t:
            s += 1
        return s

    def _skor_rows(self, rows) -> int:
        score = 0
        for r in rows[:40]:
            cells = [str(c).strip() for c in r]
            if self._is_header(cells):
                score += 10 + max(0, len([c for c in cells if c]) - 2)
            if cells and self._sel_nama(cells):
                score += 2
            # banyak kolom non-kosong = delimiter masuk akal
            nn = sum(1 for c in cells if c)
            if nn >= 4:
                score += 1
        return score

    def _is_header(self, cells) -> bool:
        norms = [re.sub(r"\s+", " ", c).casefold().strip(":'\" ") for c in cells if c]
        if not norms:
            return False
        has_tgl = any(any(h == n or h in n for h in self._HDR_TGL) for n in norms)
        has_saldo = any(any(h == n or h in n for h in self._HDR_SALDO) for n in norms)
        return has_tgl and has_saldo

    def _sel_nama(self, cells) -> bool:
        c0 = re.sub(r"\s+", " ", str(cells[0])).casefold().strip(":'\" ")
        return c0 in ("nama", "name", "nama/name", "account name")

    def _ambil_owner(self, cells) -> str:
        for c in cells[1:]:
            s = str(c).lstrip("'").strip()
            if s and s not in ("=", ":", "-"):
                return s
        return ""

    def _norm_row(self, d: dict) -> dict:
        """Map alias header → Tanggal/Keterangan/Jumlah/Saldo."""
        key_map = {}
        for k in d:
            nk = re.sub(r"\s+", " ", str(k)).casefold().strip(":'\" ")
            if any(h == nk or h in nk for h in self._HDR_TGL):
                key_map[k] = "Tanggal"
            elif any(h == nk or h in nk for h in self._HDR_KET):
                key_map[k] = "Keterangan"
            elif any(h == nk or h in nk for h in self._HDR_JUMLAH):
                key_map[k] = "Jumlah"
            elif any(h == nk or h in nk for h in self._HDR_SALDO):
                key_map[k] = "Saldo"
        out = dict(d)
        for old, new in key_map.items():
            if new not in out or not out.get(new):
                out[new] = d.get(old, "")
        return out


class MandiriParser(BaseParser):
    source_key = "bank"

    def parse(self, path, flow=""):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        allrows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        hidx = None
        for i, r in enumerate(allrows):
            cells = [str(c).strip() if c is not None else "" for c in r]
            # Header e-Statement: 'Nama/Name | : | SITI NURUL WIRDAH | ...'
            if not self.meta.get("owner_name") and any(c.startswith("Nama/Name") for c in cells):
                j = next(k for k, c in enumerate(cells) if c.startswith("Nama/Name"))
                owner = next((c for c in cells[j + 1:] if c and c != ":"), "")
                if owner:
                    self.meta["owner_name"] = owner
            if "Tanggal" in cells and "Keterangan" in cells:
                hidx = i
                break
        if hidx is None:
            return []
        hdr = [str(c).strip() if c is not None else "" for c in allrows[hidx]]

        def colof(name):
            return hdr.index(name) if name in hdr else None

        c_no, c_tgl, c_ket = colof("No"), colof("Tanggal"), colof("Keterangan")
        c_in, c_out, c_saldo = (
            colof("Dana Masuk (IDR)"),
            colof("Dana Keluar (IDR)"),
            colof("Saldo (IDR)"),
        )

        def cell(r, j):
            return r[j] if (j is not None and j < len(r)) else None

        out = []
        i, n = hidx + 1, len(allrows)
        while i < n:
            r = allrows[i]
            cells = [str(c).strip() if c is not None else "" for c in r]
            if "Date" in cells and "Remarks" in cells:  # sub-header bahasa Inggris
                i += 1
                continue
            no, tgl = cell(r, c_no), cell(r, c_tgl)
            if (no in (None, "")) and (tgl in (None, "")):
                i += 1
                continue

            datestr = str(tgl).strip() if tgl else ""
            ket = str(cell(r, c_ket) or "").strip()
            masuk = parse_decimal(cell(r, c_in), "id")
            keluar = parse_decimal(cell(r, c_out), "id")
            saldo = parse_decimal(cell(r, c_saldo), "id")

            timestr = ""
            if i + 1 < n:
                nr = allrows[i + 1]
                ntgl = str(cell(nr, c_tgl) or "").strip()
                nno = cell(nr, c_no)
                if (nno in (None, "")) and re.search(r"\d{1,2}:\d{2}", ntgl):
                    timestr = ntgl.replace("WIB", "").strip()
                    nket = str(cell(nr, c_ket) or "").strip()
                    if nket:
                        ket = f"{ket} {nket}".strip()
                    i += 1

            occurred = parse_dt(f"{datestr} {timestr}".strip(), dayfirst=True)
            money = masuk - keluar
            row = {
                "source_type": "bank",
                "occurred_at": occurred,
                "posted_date": occurred.date() if occurred else None,
                "jenis": "admin"
                if (money < 0 and is_admin_fee("mandiri", ket, abs(money)))
                else _jenis_from_money(money),
                "amount": abs(money),
                "credit_delta": Decimal("0"),
                "money_delta": money,
                "fee": Decimal("0"),
                "bonus": Decimal("0"),
                "balance_after": saldo,
                "ticket_no": "",
                "username": "",
                "reference": "",
                "counterparty": extract_mandiri_name(ket.replace("\n", " ")),
                "description": ket.replace("\n", " "),
                "raw": {"Tanggal": datestr, "Jam": timestr, "Keterangan": ket.replace("\n", " "),
                        "Masuk": str(masuk), "Keluar": str(keluar), "Saldo": str(saldo)},
            }
            row["row_hash"] = row_hash("mandiri", [saldo, occurred, money, ket[:30]])
            out.append(row)
            i += 1
        return out

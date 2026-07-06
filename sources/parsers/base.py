"""Helper bersama untuk semua parser sumber.

Tiap parser meng-output list of dict (baris kanonik) dengan kunci yang sama
dengan field model `transactions.Transaction`.
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from dateutil import parser as dateparser

# Ticket panel: D/W + 6-9 digit (mis. D1757153, W1757092). Sengaja dibatasi agar
# TIDAK ikut menangkap reference gateway yang panjang (>=12 digit).
TICKET_RE = re.compile(r"\b([DW]\d{6,9})\b")
# Reference gateway: 1 huruf + >=12 digit (mis. F260627206100206205).
REF_RE = re.compile(r"\b([A-Z]\d{12,})\b")
# Non-alfabet (angka/simbol) pada nama -> diganti spasi sebelum fuzzy matching.
NAME_CLEAN_RE = re.compile(r"[^A-Za-z\s]")


def clean_name(text):
    """Nama utk matching: buang non-alfabet (angka/simbol -> spasi), rapikan spasi.
    'John123 Smith!' -> 'John Smith'; 'John123Smith' -> 'John Smith'."""
    s = NAME_CLEAN_RE.sub(" ", str(text or ""))
    return re.sub(r"\s+", " ", s).strip()


# --- Reader mentah xlsx (tahan exporter non-standar tanpa <dimension>/styles) ---
def _xlsx_local(tag):
    return tag.rsplit("}", 1)[-1]


def _xlsx_col_idx(ref):
    m = re.match(r"([A-Z]+)", ref or "")
    if not m:
        return 0
    n = 0
    for ch in m.group(1):
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _raw_xlsx_rows(path, nrows=None):
    """Baca xlsx via zip+xml langsung (abaikan styles.xml). -> list[list[str]]."""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        sheet = "xl/worksheets/sheet1.xml"
        if sheet not in names:
            cand = sorted(n for n in names
                          if n.startswith("xl/worksheets/") and n.endswith(".xml"))
            sheet = cand[0] if cand else None
        sst = []
        if "xl/sharedStrings.xml" in names:
            for _, el in ET.iterparse(io.BytesIO(z.read("xl/sharedStrings.xml"))):
                if _xlsx_local(el.tag) == "si":
                    sst.append("".join(t.text or "" for t in el.iter()
                                       if _xlsx_local(t.tag) == "t"))
                    el.clear()
        rows = []
        for _, el in ET.iterparse(io.BytesIO(z.read(sheet))):
            if _xlsx_local(el.tag) != "row":
                continue
            cells, maxc = {}, 0
            for c in el:
                if _xlsx_local(c.tag) != "c":
                    continue
                idx = _xlsx_col_idx(c.attrib.get("r", "")); t = c.attrib.get("t", "")
                v, vt, ist = "", None, None
                for ch in c:
                    lt = _xlsx_local(ch.tag)
                    if lt == "v":
                        vt = ch.text
                    elif lt == "is":
                        ist = "".join(x.text or "" for x in ch.iter()
                                      if _xlsx_local(x.tag) == "t")
                if t == "s" and vt is not None:
                    try:
                        v = sst[int(vt)]
                    except (ValueError, IndexError):
                        v = vt
                elif t == "inlineStr" and ist is not None:
                    v = ist
                elif vt is not None:
                    v = vt
                cells[idx] = v; maxc = max(maxc, idx)
            rows.append([cells.get(i, "") for i in range(maxc + 1)])
            el.clear()
            if nrows is not None and len(rows) >= nrows:
                break
        return rows


def read_xlsx_rows(path, header_row=1, sheet=None):
    """Baca xlsx -> (headers, list_of_dict). header_row 1-based.
    Fallback ke reader mentah bila openpyxl gagal / hanya kebaca <= header_row baris
    (exporter non-standar tanpa <dimension>)."""
    grid = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    except Exception:
        grid = None
    if grid is None:
        grid = _raw_xlsx_rows(path)  # openpyxl gagal total (mis. styles.xml rusak)
    elif len(grid) <= header_row:
        # openpyxl sukses tapi <= baris header (0 baris data). Bisa file well-formed
        # yang memang tanpa data, ATAU exporter non-standar (tanpa <dimension>) yang
        # terbaca 0 baris. Pakai raw HANYA bila menemukan LEBIH banyak baris — jangan
        # buang hasil openpyxl (nilai typed float/tanggal) untuk file well-formed.
        raw = _raw_xlsx_rows(path)
        if len(raw) > len(grid):
            grid = raw
    headers, out = None, []
    for i, row in enumerate(grid, start=1):
        if i < header_row:
            continue
        if i == header_row:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if row is None or all(c is None or c == "" for c in row):
            continue
        out.append({h: c for h, c in zip(headers, row) if h})
    return headers, (out or [])


def parse_decimal(value, number_format="intl"):
    """Parse angka (str/float/int) -> Decimal. Mendukung format intl & ID."""
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace("Rp", "").replace("'", "").strip()
    if not s:
        return Decimal("0")
    neg = s.startswith("-") or s.endswith("DB") or s.endswith("Db")
    s = s.replace("DB", "").replace("Db", "").replace("CR", "").replace("Cr", "")
    s = s.strip().lstrip("+-").strip()
    if number_format == "id":  # 1.000,00 -> 1000.00
        s = s.replace(".", "").replace(",", ".")
    else:  # intl 1,000.00 -> 1000.00
        s = s.replace(",", "")
    s = re.sub(r"[^0-9.]", "", s)
    if s in ("", "."):
        return Decimal("0")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return Decimal("0")
    return -d if neg else d


def parse_dt(value, dayfirst=False):
    """Parse tanggal/jam -> datetime (atau None)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip().lstrip("'")
    if not s:
        return None
    try:
        return dateparser.parse(s, dayfirst=dayfirst)
    except (ValueError, OverflowError, TypeError):
        return None


def extract_ticket(text):
    m = TICKET_RE.search(text or "")
    return m.group(1) if m else ""


def extract_ref(text):
    m = REF_RE.search(text or "")
    return m.group(1) if m else ""


def row_hash(source_key, parts):
    """Hash stabil untuk idempotensi (cegah re-import baris yang sama)."""
    h = hashlib.sha256()
    h.update(source_key.encode())
    for p in parts:
        h.update(b"|")
        h.update(str(p).encode())
    return h.hexdigest()


def read_csv_raw(path, encoding="utf-8-sig", delimiter=","):
    """Baca CSV mentah -> list of list (semua baris, termasuk preamble)."""
    with open(path, newline="", encoding=encoding, errors="replace") as f:
        return list(csv.reader(f, delimiter=delimiter))


def rows_to_dicts(rows, header_idx):
    """list-of-list -> list-of-dict pakai baris header ke-`header_idx` (0-based)."""
    headers = [str(c).strip() for c in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1 :]:
        if not any(str(c).strip() for c in r):
            continue
        out.append({(h if h else f"col{i}"): (r[i] if i < len(r) else "") for i, h in enumerate(headers)})
    return headers, out


def first_part(value, sep=","):
    """Ambil bagian sebelum `sep` (mis. buang ', Platform: ...' di tanggal Panel)."""
    return str(value or "").split(sep)[0].strip()


def bank_code(value, sep="|"):
    """Kode bank/dompet = segmen pertama sebelum `sep`, huruf besar. '' bila kosong.

    Contoh: "DANA|Mhd Ilyas|0822" -> "DANA"; "BCA 7126201591" (sep=" ") -> "BCA".
    """
    return first_part(value, sep).upper()


# Sumber -> (key player_bank, sep, key bank_title, sep). Dipakai parser & backfill.
_BANK_FIELDS = {
    "panel": ("Player Bank", "|", "Bank Title", "|"),
    "bracket": ("No. Rek Bank Member", " ", "Bank", "|"),
}


def derive_bank_fields(source_key, raw):
    """(player_bank, bank_title) dari `raw` sesuai sumber. ('', '') bila tak berlaku.

    Sisi kredit (panel/bracket) saja; sumber uang (bank/gateway) -> ('', '').
    """
    spec = _BANK_FIELDS.get(source_key)
    if not spec:
        return "", ""
    raw = raw or {}
    pkey, psep, tkey, tsep = spec
    return bank_code(raw.get(pkey), psep), bank_code(raw.get(tkey), tsep)


def parse_bank_triplet(value):
    """String COR "KODE - NOREK - NAMA" -> (kode, norek, nama). Nama boleh memuat
    ' - ' internal (mis. '.../ WITHDRAW BCA') -> hanya split 2 pemisah pertama."""
    s = str(value or "").strip()
    if not s:
        return "", "", ""
    parts = [p.strip() for p in s.split(" - ", 2)]
    while len(parts) < 3:
        parts.append("")
    return parts[0].upper(), parts[1], parts[2]


class BaseParser:
    """Interface parser. Subclass set `source_key` & implement `parse`."""

    source_key: str | None = None

    def parse(self, path, flow="") -> list[dict]:
        raise NotImplementedError

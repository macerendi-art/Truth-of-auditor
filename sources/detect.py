"""Auto-deteksi jenis sumber file dari tanda-tangan header/isi.

detect_source(path, filename) -> [{"parser_key", "confidence"}] terurut menurun.
parser_key mengacu ke kunci sources.services.PARSERS.
"""
import os

import openpyxl


def _ext(filename):
    return os.path.splitext(filename)[1].lower()


def _xlsx_tokens(path, max_rows=3):
    from .parsers.base import _raw_xlsx_rows
    toks = set()
    grid = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [list(r) for _, r in zip(range(max_rows), ws.iter_rows(values_only=True))]
        wb.close()
    except Exception:
        grid = None
    if not grid:
        try:
            grid = _raw_xlsx_rows(path)[:max_rows]
        except Exception:
            grid = []
    for row in grid:
        for c in row:
            if c is not None and c != "":
                toks.add(str(c).strip().lower())
    return toks


def _csv_text(path, n=6):
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            return "\n".join(line.strip().lower() for _, line in zip(range(n), f))
    except Exception:
        return ""


def _has(toks, needle):
    return any(needle in t for t in toks)


def detect_source(path, filename=""):
    filename = filename or os.path.basename(path)
    ext = _ext(filename)
    fn = filename.lower()
    scored = {}

    def add(key, conf):
        scored[key] = max(scored.get(key, 0.0), conf)

    if ext in (".xlsx", ".xls"):
        t = _xlsx_tokens(path)
        if _has(t, "ticket number") and _has(t, "user name") and (_has(t, "deposit amount") or _has(t, "withdrawal amount")):
            add("panel", 0.95)
        if _has(t, "ticket number") and (_has(t, "admin fee") or _has(t, "account title")) and not _has(t, "deposit amount"):
            add("nxpay", 0.90)
        if _has(t, "kategori") and (_has(t, "credit awal") or _has(t, "credit akhir")):
            add("bracket", 0.95)
        if _has(t, "client reference") and (_has(t, "settlement time") or _has(t, "txn id")):
            add("qrflyer", 0.90)
        if _has(t, "qris") or _has(t, "qr flyer") or "qrflyer" in fn or "qris" in fn or "qr flyer" in fn:
            add("qrflyer", 0.85)
        if _has(t, "e-statement") or _has(t, "rekening koran") or "mandiri" in fn:
            add("mandiri", 0.80)
        if _has(t, "orderid") and _has(t, "grandtotal") and _has(t, "branchnominal"):
            add("cor_qris_gateway", 0.95)
        if _has(t, "whitelabel transaction id") and _has(t, "nmid"):
            add("qhoki", 0.95)
        if _has(t, "from bank") and _has(t, "destination bank") and _has(t, "approved date"):
            add("cor_panel_bank", 0.95)
        if _has(t, "transaction id") and _has(t, "amount") and _has(t, "bonus") \
                and not _has(t, "kategori"):
            add("cor_panel_qris", 0.90)
    elif ext == ".csv":
        c = _csv_text(path)
        if "mutasi_debet" in c or "mutasi_kredit" in c or "tgl_tran" in c:
            add("bri", 0.95)
        if ("cabang" in c and "keterangan" in c and "saldo" in c) or "bca" in fn:
            add("bca_csv", 0.85)
    elif ext == ".pdf":
        add("bca_pdf", 0.75)

    ranked = sorted(scored.items(), key=lambda kv: kv[1], reverse=True)
    return [{"parser_key": k, "confidence": c} for k, c in ranked]

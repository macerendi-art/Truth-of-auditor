"""Bonus masuk berkas ekspor (Fase 6).

(A) `bonus_sheet` — SATU tabel untuk semua ember (Cocok / Hanya Panel / Hanya
    Bracket / Agregat), dibedakan kolom Status. Baris agregat muat di kolom yang
    sama: Username kosong, Kategori = kategori bracket, kedua kolom nominal =
    total masing-masing sisi.
(B) Workbook batch mendapat sheet itu query-time — HANYA bila tanggal batch
    punya baris bonus, sehingga batch tanpa data bonus tetap identik format lama.
(C) View `/bonus/export/` menghormati `?dari=&sampai=&kategori=` yang sedang
    aktif. **Invarian utama:** berkasnya memuat PERSIS baris yang terlihat di
    halaman untuk query yang sama — karena itu kedua view mem-parse filter lewat
    `_bonus_params`, bukan dua salinan logika. Ekspor yang menyimpang dari layar
    lebih buruk daripada tidak ada ekspor: operator memakainya sebagai bukti.
"""
import io
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.models import AuditLog
from reconciliation.models import MatchRun, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction
from web.bonus import rekonsiliasi_bonus
from web.exports import XLSX_CT, bonus_sheet, build_batch_workbook
from web.tests_bonus import TGL, _BonusData
from web.tests_bonus_agregat import D3, D4, KAT_ROLL_BRACKET, KAT_ROLL_PANEL, _AgregatData
from web.views import REL_LABELS

User = get_user_model()

HEADER = ["Tanggal", "Username", "Kategori", "Nominal Panel",
          "Nominal Bracket", "Selisih", "Status"]


def _header_index(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Tanggal":
            return i
    raise AssertionError("baris header 'Tanggal' tidak ditemukan")


def _baris_data(ws):
    """Baris isi tabel (tanpa caption, header, maupun baris TOTAL)."""
    out = []
    for row in list(ws.iter_rows(values_only=True))[_header_index(ws):]:
        if row and row[0] == "TOTAL":
            break
        out.append(row)
    return out


def _baris_total(ws):
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "TOTAL":
            return row
    raise AssertionError("baris TOTAL tidak ditemukan")


def _sheet(data, label=""):
    wb = Workbook()
    wb.remove(wb.active)
    return bonus_sheet(wb, data, "Rekonsiliasi Bonus", label)


# --------------------------------------------------------------------------- #
#  (A) bonus_sheet — unit, tanpa view                                         #
# --------------------------------------------------------------------------- #
class BonusSheetTests(_AgregatData):
    def test_header_dan_caption(self):
        self.panel_row("cici", "40000", tanggal=D4)
        ws = _sheet(self.recon(), "03/08/2026 – 04/08/2026")
        self.assertEqual(ws["A1"].value, "03/08/2026 – 04/08/2026")
        self.assertEqual([c.value for c in ws[_header_index(ws)]], HEADER)

    def test_tiga_ember_lama_dengan_status_dan_tanda_selisih(self):
        self.panel_row("cici", "40000", tanggal=D4)
        self.bracket_row("cici", "40000", tanggal=D4)
        self.panel_row("sendy", "20000", tanggal=D4)
        self.bracket_row("andi", "15000", tanggal=D4)
        baris = {r[6]: r for r in _baris_data(_sheet(self.recon()))}
        self.assertEqual(set(baris), {"Cocok", "Hanya Panel", "Hanya Bracket"})
        # Selisih = Nominal Panel − Nominal Bracket di SEMUA baris.
        self.assertEqual(baris["Cocok"][3:7],
                         (40000.0, 40000.0, 0.0, "Cocok"))
        self.assertEqual(baris["Hanya Panel"][3:7],
                         (20000.0, "", 20000.0, "Hanya Panel"))
        self.assertEqual(baris["Hanya Bracket"][3:7],
                         ("", 15000.0, -15000.0, "Hanya Bracket"))

    def test_angka_masuk_sebagai_angka(self):
        self.panel_row("cici", "40000", tanggal=D4)
        r = _baris_data(_sheet(self.recon()))[0]
        self.assertIsInstance(r[3], float)

    def test_kolom_identitas_baris_biasa(self):
        self.panel_row("Cici", "40000", tanggal=D4, kategori="Lucky Draw")
        r = _baris_data(_sheet(self.recon()))[0]
        self.assertEqual(r[0], "04/08/2026")
        self.assertEqual(r[1], "Cici")
        self.assertEqual(r[2], "Lucky Draw")

    def test_baris_agregat_muat_di_kolom_yang_sama(self):
        self.panel_cor("a", "500000.00", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.panel_cor("b", "538747.20", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.lump("1038747.00", KAT_ROLL_BRACKET)
        r = _baris_data(_sheet(self.recon()))[0]
        self.assertEqual(r[0], "04/08/2026")
        self.assertEqual(r[1], "")                     # username kosong
        self.assertEqual(r[2], KAT_ROLL_BRACKET)       # kategori BRACKET
        self.assertEqual(r[3], 1038747.20)
        self.assertEqual(r[4], 1038747.00)
        self.assertAlmostEqual(r[5], 0.20, places=2)
        self.assertEqual(r[6], "Agregat")

    def test_agregat_berselisih_ditandai(self):
        self.panel_cor("a", "157500", kategori="Single Deposit", tanggal=D4)
        self.lump("215000", "SINGLE DEPOSIT")
        r = _baris_data(_sheet(self.recon()))[0]
        self.assertEqual(r[5], -57500.0)
        self.assertEqual(r[6], "Agregat (selisih)")

    def test_tanggal_efektif_dipakai_bukan_tanggal_pembukuan(self):
        self.panel_cor("a", "92925", kategori="Daily Login", tanggal=D3)
        self.lump("92925", "DAILY LOGIN", tanggal=D4,
                  desc="LOGIN & SPIN (DAILY SPIN BONUS) TGL 03.08.2026\nPlayer:")
        self.assertEqual(_baris_data(_sheet(self.recon()))[0][0], "03/08/2026")

    def test_total_menjumlah_semua_ember(self):
        self.panel_row("cici", "40000", tanggal=D4)
        self.bracket_row("cici", "40000", tanggal=D4)
        self.panel_row("sendy", "20000", tanggal=D4)
        self.bracket_row("andi", "15000", tanggal=D4)
        self.panel_cor("a", "100000", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.lump("100000", KAT_ROLL_BRACKET)
        total = _baris_total(_sheet(self.recon()))
        self.assertEqual(total[3], 160000.0)   # 40.000 + 20.000 + 100.000
        self.assertEqual(total[4], 155000.0)   # 40.000 + 15.000 + 100.000
        self.assertEqual(total[5], 5000.0)

    def test_nexus_tanpa_kunci_agregat_tetap_terekspor(self):
        # `data` toko Nexus tak punya kunci "agregat" sama sekali — sheet wajib
        # tetap terbentuk (dibaca lewat .get, bukan indeks langsung).
        self.panel_row("cici", "40000", tanggal=D4)
        data = self.recon()
        self.assertNotIn("agregat", data)
        self.assertEqual(len(_baris_data(_sheet(data))), 1)


# --------------------------------------------------------------------------- #
#  (B) build_batch_workbook diperkaya                                         #
# --------------------------------------------------------------------------- #
class BatchWorkbookBonusTests(_BonusData):
    def setUp(self):
        super().setUp()
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1})[0]

    def _batch(self, recon_date=TGL):
        batch = ReconBatch.objects.create(
            toko=self.toko, tolerance=self.tol, recon_date=recon_date,
            summary={"dp": {}, "wd": {}, "buckets": {}})
        MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol,
            batch=batch, summary={})
        return batch

    def test_sheet_bonus_ditambahkan_bila_ada_baris(self):
        self.panel_row("cici", "40000")
        self.bracket_row("cici", "40000")
        wb = build_batch_workbook(self._batch(), 1, REL_LABELS)
        self.assertIn("Rekonsiliasi Bonus", wb.sheetnames)

    def test_tanpa_baris_bonus_workbook_identik_format_lama(self):
        wb = build_batch_workbook(self._batch(), 1, REL_LABELS)
        self.assertNotIn("Rekonsiliasi Bonus", wb.sheetnames)

    def test_bonus_tanggal_lain_tak_bikin_sheet(self):
        self.panel_row("cici", "40000", tanggal=date(2026, 7, 1))
        wb = build_batch_workbook(self._batch(recon_date=TGL), 1, REL_LABELS)
        self.assertNotIn("Rekonsiliasi Bonus", wb.sheetnames)

    def test_hanya_bracket_pun_cukup(self):
        self.bracket_row("andi", "15000")
        wb = build_batch_workbook(self._batch(), 1, REL_LABELS)
        self.assertIn("Rekonsiliasi Bonus", wb.sheetnames)


# --------------------------------------------------------------------------- #
#  (C) view /bonus/export/                                                    #
# --------------------------------------------------------------------------- #
class ExportBonusViewTests(_AgregatData):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(
            "sup_bn", "s@a.co", "pw12345", role="supervisor")
        self.client.login(username="sup_bn", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})

    def _q(self, **extra):
        q = {"dari": D3.isoformat(), "sampai": D4.isoformat()}
        q.update(extra)
        return q

    def _ws(self, **extra):
        r = self.client.get(reverse("export_bonus"), self._q(**extra))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], XLSX_CT)
        return load_workbook(io.BytesIO(r.content)).active

    def test_butuh_login(self):
        self.client.logout()
        r = self.client.get(reverse("export_bonus"))
        self.assertEqual(r.status_code, 302)
        self.assertIn("login", r["Location"])

    def test_nama_file_dan_sheet(self):
        self.panel_row("cici", "40000", tanggal=D4)
        r = self.client.get(reverse("export_bonus"), self._q())
        cd = r["Content-Disposition"]
        self.assertIn("bonus_", cd)
        self.assertIn(self.toko.name.replace(" ", "_"), cd)
        self.assertIn(D3.isoformat(), cd)
        self.assertIn(D4.isoformat(), cd)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertIn("Rekonsiliasi Bonus", wb.sheetnames)

    def test_tanpa_data_tetap_xlsx_valid(self):
        # Aturan "hanya bila ada baris" berlaku untuk workbook batch; view ini
        # selalu menulis sheetnya (workbook tanpa sheet = openpyxl melempar).
        ws = self._ws()
        self.assertEqual(_baris_data(ws), [])
        self.assertEqual(_baris_total(ws)[3], 0.0)

    def test_hormati_rentang(self):
        self.panel_row("lama", "11000", tanggal=date(2026, 7, 1))
        self.panel_row("baru", "22000", tanggal=D4)
        user = {r[1] for r in _baris_data(self._ws())}
        self.assertEqual(user, {"baru"})

    def test_ekspor_persis_sama_dengan_layar(self):
        """Invarian: jumlah & isi baris berkas == baris di halaman, untuk query
        yang sama — termasuk saat filter kategori aktif."""
        self.panel_cor("a", "500000.00", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.panel_cor("b", "538747.20", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.lump("1038747.00", KAT_ROLL_BRACKET)
        self.panel_cor("solo", "40000", kategori="Manual Freebet", tanggal=D4)
        self.bracket_row("solo", "40000", kategori="BONUS HARIAN", tanggal=D4)
        self.panel_cor("yatim", "12000", kategori="Manual Freebet", tanggal=D4)

        for kategori in ("", KAT_ROLL_PANEL, "Manual Freebet"):
            with self.subTest(kategori=kategori):
                q = self._q(kategori=kategori)
                data = self.client.get(reverse("bonus_recon"), q).context["data"]
                diharapkan = (len(data["cocok"]) + len(data["panel_only"])
                              + len(data["bracket_only"])
                              + len(data.get("agregat") or []))
                r = self.client.get(reverse("export_bonus"), q)
                baris = _baris_data(load_workbook(io.BytesIO(r.content)).active)
                self.assertEqual(len(baris), diharapkan)

    def test_filter_kategori_menyingkirkan_baris_lain(self):
        self.panel_cor("a", "100000", kategori=KAT_ROLL_PANEL, tanggal=D4)
        self.lump("100000", KAT_ROLL_BRACKET)
        self.panel_cor("yatim", "12000", kategori="Manual Freebet", tanggal=D4)
        baris = _baris_data(self._ws(kategori="Manual Freebet"))
        self.assertEqual([r[1] for r in baris], ["yatim"])
        self.assertNotIn(KAT_ROLL_BRACKET, {r[2] for r in baris})

    def test_scope_rbac_toko_lain_tak_bocor(self):
        lain = Toko.objects.exclude(pk=self.toko.pk).first()
        up = Upload.objects.create(
            source_type=SourceType.objects.get(key="panel_bonus"), toko=lain)
        Transaction.objects.create(
            upload=up, source_type=up.source_type, toko=lain, jenis="bonus",
            amount=Decimal("99000"), money_delta=Decimal("0"),
            username="tokolain", posted_date=D4, description="Lucky Draw",
            raw={"Kategori": "Lucky Draw"}, row_hash="bn-lain")
        self.client.logout()
        aud = User.objects.create_user("aud_bn2", "a@a.co", "pw12345",
                                       role="auditor")
        aud.allowed_tokos.set([self.toko])
        self.client.login(username="aud_bn2", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})
        r = self.client.get(reverse("export_bonus"), self._q())
        self.assertEqual(r["Content-Type"], XLSX_CT)
        self.assertNotIn(b"tokolain", r.content)

    def test_audit_tercatat(self):
        self.client.get(reverse("export_bonus"), self._q())
        self.assertTrue(AuditLog.objects.filter(aksi="export_bonus").exists())


class ExportBonusButtonTests(_BonusData):
    def setUp(self):
        super().setUp()
        u = User.objects.create_user("sup_bn3", "s3@a.co", "pw12345",
                                     role="supervisor")
        self.client.login(username="sup_bn3", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})

    def test_tombol_membawa_filter_aktif(self):
        self.panel_row("cici", "40000")
        r = self.client.get(reverse("bonus_recon"),
                            {"dari": "2026-07-01", "sampai": "2026-07-31",
                             "kategori": "Lucky Draw"})
        html = r.content.decode()
        self.assertIn(reverse("export_bonus"), html)
        self.assertIn("dari=2026-07-01", html)
        self.assertIn("sampai=2026-07-31", html)
        self.assertIn("kategori=Lucky%20Draw", html)


class BonusParamsTests(_BonusData):
    """`_bonus_params` adalah SATU jalur parsing untuk halaman + ekspornya."""

    def test_dipakai_kedua_view(self):
        import inspect

        from web import views

        for fn in (views.bonus_recon, views.export_bonus):
            self.assertIn("_bonus_params", inspect.getsource(fn), fn.__name__)

    def test_default_tiga_puluh_hari(self):
        from django.test import RequestFactory

        from web.views import _bonus_params

        req = RequestFactory().get("/bonus/", {"sampai": "2026-08-04"})
        dari, sampai, kategori = _bonus_params(req)
        self.assertEqual(sampai, D4)
        self.assertEqual((sampai - dari).days, 30)
        self.assertEqual(kategori, "")

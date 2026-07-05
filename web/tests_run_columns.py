"""Kolom Panel dipecah (User ID / Full Name / Player Bank / Bank Title / Handler)
+ halaman run menampilkan Batch #N per-toko (bukan pk) dengan link balik ke batch."""
import io
import re
from datetime import datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from reconciliation.models import MatchResult, MatchRun, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction

User = get_user_model()


def _row_cells(html, pk):
    """Ambil isi teks tiap <td> dari baris hasil <tr id="res-<pk>">."""
    m = re.search(r'<tr id="res-%s">(.*?)</tr>' % pk, html, re.S)
    assert m, f"baris res-{pk} tidak ditemukan"
    cells = re.findall(r"<td[^>]*>(.*?)</td>", m.group(1), re.S)
    return [re.sub(r"<[^>]+>", " ", c).strip() for c in cells]


class _Base(TestCase):
    def setUp(self):
        User.objects.create_user("aud", "a@a.co", "pw12345", role="supervisor")
        self.client.login(username="aud", password="pw12345")
        self.lbs = Toko.objects.get(key="lbs")
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.bank = SourceType.objects.get_or_create(key="bank", defaults={"name": "Bank"})[0]
        self.up = Upload.objects.create(source_type=self.panel, toko=self.lbs)
        self.batch = ReconBatch.objects.create(toko=self.lbs, tolerance=self.tol)
        self.run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=self.batch
        )

    def _tx(self, st, row_hash, **kw):
        defaults = dict(
            upload=self.up, source_type=st, toko=self.lbs, jenis="depo",
            amount=Decimal("50000"), occurred_at=datetime(2026, 6, 27, 10, 0), raw={},
        )
        defaults.update(kw)
        return Transaction.objects.create(row_hash=row_hash, **defaults)


class SplitPanelColumnsTests(_Base):
    """TASK 1: kolom terpisah User ID / Full Name / Player Bank / Bank Title / Handler."""

    RAW = {
        "Player Bank": "DANA|fajar Pratama |083822153879",
        "Bank Title": "BCA|HENDI|7126201591",
        "Handler": "Mozart K25",
    }

    def test_header_kolom_baru_ada(self):
        resp = self.client.get(reverse("run_detail", args=[self.run.pk]))
        self.assertEqual(resp.status_code, 200)
        for th in ("<th>Username</th>", "<th>Nama Lengkap</th>", '<th class="col-hide">Player Bank</th>',
                   '<th class="col-hide">Bank Title</th>', '<th class="col-hide">Handler</th>'):
            self.assertContains(resp, th)

    def test_nilai_terpisah_per_sel_dan_sel_panel_hanya_ticket_tanggal(self):
        left = self._tx(
            self.panel, "rc1",
            ticket_no="D0012345", username="budi123", counterparty="BUDI SANTOSO",
            raw=self.RAW,
        )
        right = self._tx(self.bank, "rc2")
        r = MatchResult.objects.create(
            run=self.run, bucket=MatchResult.Bucket.COCOK, left=left, right=right,
        )
        resp = self.client.get(reverse("run_detail", args=[self.run.pk]))
        cells = _row_cells(resp.content.decode(), r.pk)
        # 12 kolom: Status, Panel, User ID, Full Name, Player Bank, Bank Title,
        # Handler, Amount, Kanan, Amount, Alasan, Aksi
        self.assertEqual(len(cells), 13)  # +1 checkbox bulk-review
        # Sel PANEL (indeks +1: kolom 0 = checkbox bulk): ticket + tanggal saja.
        self.assertIn("D0012345", cells[2])
        self.assertIn("27/06 10:00", cells[2])
        self.assertNotIn("budi123", cells[2])
        self.assertNotIn("BUDI SANTOSO", cells[2])
        # Masing-masing nilai di kolomnya sendiri.
        self.assertIn("budi123", cells[3])
        self.assertIn("BUDI SANTOSO", cells[4])
        self.assertIn("DANA|fajar Pratama |083822153879", cells[5])
        self.assertIn("BCA|HENDI|7126201591", cells[6])
        self.assertIn("Mozart K25", cells[7])

    def test_sisi_kiri_bracket_tanpa_field_panel_render_strip(self):
        """Relasi dgn kiri=Bracket (tanpa username/raw panel) → sel berisi — tanpa error."""
        bracket = SourceType.objects.get_or_create(key="bracket", defaults={"name": "Bracket"})[0]
        run = MatchRun.objects.create(
            relation=MatchRun.Relation.BRACKET_BANK, tolerance=self.tol, batch=self.batch
        )
        left = self._tx(bracket, "rc3", ticket_no="BR-1", username="", counterparty="")
        r = MatchResult.objects.create(
            run=run, bucket=MatchResult.Bucket.TIDAK, left=left, right=None,
        )
        resp = self.client.get(reverse("run_detail", args=[run.pk]))
        self.assertEqual(resp.status_code, 200)
        cells = _row_cells(resp.content.decode(), r.pk)
        self.assertEqual(len(cells), 13)  # +1 checkbox bulk-review
        for i in (3, 4, 5, 6, 7):  # User ID, Full Name, Player Bank, Bank Title, Handler
            self.assertEqual(cells[i], "—")

    def test_empty_state_colspan_12(self):
        resp = self.client.get(reverse("run_detail", args=[self.run.pk]))
        self.assertContains(resp, 'colspan="13"')


class RunBatchLabelTests(TestCase):
    """TASK 3: halaman run menampilkan Batch #N (nomor urut per-toko) + link ke batch."""

    def setUp(self):
        User.objects.create_user("aud", "a@a.co", "pw12345", role="supervisor")
        self.client.login(username="aud", password="pw12345")
        self.lbs = Toko.objects.get(key="lbs")
        self.lain = Toko.objects.exclude(key="lbs").first()
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        # Batch toko lain dibuat DULU → pk batch lbs != nomor urut per-toko.
        ReconBatch.objects.create(toko=self.lain, tolerance=self.tol)
        self.batch = ReconBatch.objects.create(toko=self.lbs, tolerance=self.tol)
        self.run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=self.batch
        )

    def test_run_page_tampilkan_nomor_batch_per_toko_dengan_link(self):
        self.assertNotEqual(self.batch.pk, 1)  # pastikan pk != nomor tampilan
        resp = self.client.get(reverse("run_detail", args=[self.run.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Batch #1")  # nomor urut per-toko, bukan pk
        self.assertNotContains(resp, f"Batch #{self.batch.pk}")
        self.assertContains(resp, f'href="{reverse("batch_detail", args=[self.batch.pk])}"')
        self.assertContains(resp, f"(Run #{self.run.pk})")

    def test_run_tanpa_batch_template_aman_tanpa_label_batch(self):
        """Run lama (batch=None): template render tanpa error & tanpa 'Batch #'."""
        run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=None
        )
        html = render_to_string("web/run_detail.html", {
            "run": run, "page": Paginator(MatchResult.objects.none(), 40).get_page(1),
            "bucket": "", "left_label": "Panel", "right_label": "Bank/Gateway",
            "batch": None, "batch_no": None,
        })
        self.assertNotIn("Batch #", html)
        self.assertIn(f"(Run #{run.pk})", html)

    def test_run_tanpa_batch_view_404_bukan_500(self):
        """Akses view utk run batch=None: filter RBAC batch__toko → 404 (bukan crash)."""
        run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=None
        )
        resp = self.client.get(reverse("run_detail", args=[run.pk]))
        self.assertEqual(resp.status_code, 404)


class ExportRunPanelColumnsTests(_Base):
    """Export Excel meniru tabel layar: kolom Player Bank / Bank Title / Handler sisi kiri."""

    def test_export_header_dan_nilai_kolom_panel(self):
        left = self._tx(
            self.panel, "rc9",
            ticket_no="D0012345", username="budi123", counterparty="BUDI SANTOSO",
            raw=SplitPanelColumnsTests.RAW,
        )
        right = self._tx(self.bank, "rc10")
        MatchResult.objects.create(
            run=self.run, bucket=MatchResult.Bucket.COCOK, left=left, right=right,
        )
        resp = self.client.get(reverse("export_run", args=[self.run.pk]))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Hasil"]
        headers = [c.value for c in ws[1]]
        for h in ("Panel Username", "Panel Nama Lengkap", "Panel Player Bank",
                  "Panel Bank Title", "Panel Handler"):
            self.assertIn(h, headers)
        # Nominal tidak duplikat di sisi kiri.
        self.assertEqual(headers.count("Panel Nominal"), 1)
        row2 = [c.value for c in ws[2]]
        self.assertEqual(row2[headers.index("Panel Player Bank")],
                         "DANA|fajar Pratama |083822153879")
        self.assertEqual(row2[headers.index("Panel Bank Title")], "BCA|HENDI|7126201591")
        self.assertEqual(row2[headers.index("Panel Handler")], "Mozart K25")
        self.assertEqual(row2[headers.index("Panel Username")], "budi123")

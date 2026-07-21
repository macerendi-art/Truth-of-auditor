import os
import re
import zipfile
from datetime import date as date_cls, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth import logout as auth_logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import BooleanField, Count, Exists, ExpressionWrapper, Max, Min, OuterRef, Q, Subquery, Sum
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.html import format_html, format_html_join
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from reconciliation.engine import (
    MATCHERS,
    _panel_dates,
    check_completeness,
    pending_settlement_count,
    refresh_batch_summary,
    run_batch,
    run_batches_auto,
    run_match,
)
from core.audit import catat
from core.models import AuditLog
from reconciliation.models import MatchResult, MatchRun, ReconBatch, ReviewAction, ToleranceProfile
from sources.detect import detect_source
from sources.management.commands.ingest import detect_flow
from sources.models import SourceType, Upload
from sources.services import PARSERS, ingest, is_encrypted_xlsx
from transactions.models import Transaction, specific_source_label
from web.access import is_admin, tokos_for
from web.biaya import rincian_biaya as hitung_rincian_biaya
from web.bonus import rekonsiliasi_bonus as hitung_rekonsiliasi_bonus
from web.breakdown import bracket_breakdown as hitung_bracket_breakdown, KATEGORI_KANONIK
from web.channels import breakdown_metode
from web.forms import GantiPasswordForm
from web.hutang import hutang_piutang as hitung_hutang_piutang
from web.models import FRKoreksi
from web.monthly import monthly_summary
from web.rekening import rekening_breakdown as hitung_rekening_breakdown
from web.settlement import pending_settlement_rows
from web.templatetags.web_extras import reason_label

BUCKET_META = {
    "cocok": {"label": "Cocok", "cls": "ok"},
    "perlu_tinjau": {"label": "Perlu Ditinjau", "cls": "warn"},
    "tidak_cocok": {"label": "Tidak Cocok", "cls": "bad"},
}

TX_EXPORT_LIMIT = 100_000

REL_LABELS = {
    MatchRun.Relation.PANEL_BRACKET.value: ("Panel", "Bracket"),
    MatchRun.Relation.PANEL_BANK.value: ("Panel", "Mutasi Bank"),
    MatchRun.Relation.BRACKET_BANK.value: ("Bracket", "Mutasi Bank"),
    MatchRun.Relation.SALDO.value: ("Kiri", "Kanan"),
}

# Label kolom nominal per relasi: kiri selalu kredit; kanan = Saldo Bank utk relasi
# uang, tetap Kredit/Koin utk panel<->bracket.
REL_AMOUNT_LABELS = {
    MatchRun.Relation.PANEL_BRACKET.value: ("Kredit/Koin", "Kredit/Koin"),
    MatchRun.Relation.PANEL_BANK.value: ("Kredit/Koin", "Saldo Bank"),
    MatchRun.Relation.BRACKET_BANK.value: ("Kredit/Koin", "Saldo Bank"),
    MatchRun.Relation.SALDO.value: ("Nominal", "Nominal"),
}


def _apply_sort(request, qs, allowed, default_order, default_active=None):
    """Sort server-side ber-whitelist. `allowed`={ui_key: orm_field}.
    `default_order`=list field ORM saat sort tak valid. `default_active`=(ui_key,dir)
    untuk menandai kolom default aktif. Return (qs, sort_key, direction)."""
    sort = request.GET.get("sort", "")
    direction = request.GET.get("dir", "")
    if sort not in allowed:
        if default_active and default_active[0] in allowed:
            sort, direction = default_active
        else:
            return qs.order_by(*default_order), "", ""
    if direction not in ("asc", "desc"):
        direction = "asc"
    prefix = "" if direction == "asc" else "-"
    return qs.order_by(f"{prefix}{allowed[sort]}", "id"), sort, direction


def csrf_failure(request, reason=""):
    """Token CSRF basi (tab lama / setelah redeploy) — jangan 403 mentah.

    Logout: risiko CSRF-nya sepele (paling banter dipaksa keluar), jadi
    selesaikan saja logout-nya. Selain itu: halaman ramah + link masuk.
    """
    if request.path == reverse("logout"):
        auth_logout(request)
        return redirect("login")
    return render(request, "web/csrf_failure.html", status=403)


def _active_toko(request):
    allowed = tokos_for(request.user)
    tid = request.session.get("active_toko_id")
    t = allowed.filter(id=tid).first() if tid else None
    return t or allowed.first()


@login_required
def set_toko(request):
    if request.method == "POST":
        tid = request.POST.get("toko_id", "")
        if tid.isdecimal() and tokos_for(request.user).filter(id=tid).exists():
            request.session["active_toko_id"] = int(tid)
    nxt = request.POST.get("next")
    if nxt and url_has_allowed_host_and_scheme(
        nxt, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return redirect(nxt)
    return redirect("dashboard")


@login_required
def ganti_password(request):
    """Halaman wajib ganti password (login pertama / setelah reset admin).

    @login_required: bila sesi kedaluwarsa saat di sini, submit terlempar ke
    /login/?next=... — menutup edge case sesi expired.
    """
    if request.method == "POST":
        form = GantiPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()  # set_password + simpan
            user.must_change_password = False
            user.save(update_fields=["must_change_password"])
            update_session_auth_hash(request, user)  # jaga sesi, jangan ter-logout
            catat(user, "ganti_password", user.username)
            messages.success(request, "Password berhasil diganti. Selamat bekerja!")
            return redirect("dashboard")
    else:
        form = GantiPasswordForm(user=request.user)
    return render(request, "registration/ganti_password.html", {"form": form})


@login_required
def dashboard(request):
    """Kokpit harian auditor: status hari, kalender rekon, tren selisih,
    daftar kerja — bukan sekadar statistik."""
    from datetime import timedelta

    from reconciliation.engine import check_completeness, pending_settlement_count

    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    tx = Transaction.objects.filter(toko=active)
    uploads = Upload.objects.filter(toko=active)
    runs = MatchRun.objects.filter(batch__toko=active)
    by_source = list(
        tx.values("source_type__name", "source_type__key")
        .annotate(n=Count("id"))
        .order_by("-n")
    )

    batches = list(
        ReconBatch.objects.filter(toko=active, recon_date__isnull=False)
        .order_by("recon_date")
    )
    by_date = {b.recon_date: b for b in batches}
    total_b = ReconBatch.objects.filter(toko=active).count()

    def selisih(b):
        s = b.summary or {}
        dp = abs((s.get("dp") or {}).get("selisih") or 0)
        wd = abs((s.get("wd") or {}).get("selisih") or 0)
        return dp + wd

    # --- kalender 14 hari terakhir (anchor: recon terakhir atau hari ini) ---
    today = date_cls.today()
    anchor = max(batches[-1].recon_date, today) if batches else today
    kal = []
    for i in range(13, -1, -1):
        d = anchor - timedelta(days=i)
        b = by_date.get(d)
        if b is None:
            st = ""
        else:
            tot = selisih(b)
            st = "ok" if tot == 0 else ("warn" if tot < 10_000_000 else "bad")
        kal.append({
            "d": d, "batch": b, "st": st, "today": d == today,
            "no": (ReconBatch.objects.filter(toko=active, id__lte=b.id).count() if b else None),
        })

    # --- tren selisih 30 hari kalender terakhir (bar DP/WD + garis total) ---
    tren_cutoff = anchor - timedelta(days=29)
    tren_src = [b for b in batches if b.recon_date and b.recon_date >= tren_cutoff]
    mx = max((selisih(b) for b in tren_src), default=0) or 1
    tren = []
    for b in tren_src:
        s = b.summary or {}
        dp = abs((s.get("dp") or {}).get("selisih") or 0)
        wd = abs((s.get("wd") or {}).get("selisih") or 0)
        tren.append({
            "b": b, "dp": dp, "wd": wd,
            "hdp": round(100 * dp / mx), "hwd": round(100 * wd / mx),
            "tot": dp + wd, "htot": round(100 * (dp + wd) / mx),
        })

    # --- kartu status ---
    last = batches[-1] if batches else None

    # --- ringkasan panel (trx & nilai DP/WD) batch terakhir ---
    # consumed_by_batch=last: baris panel yang terkunci ke batch itulah potret
    # harinya (termasuk baris retro yang ditulis-balik) — konsisten dengan
    # summary["dp"]["panel"].
    panel_sum = None
    metode = None
    if last is not None:
        _pr = Transaction.objects.filter(
            consumed_by_batch=last, source_type__key="panel", is_duplicate=False,
        )
        _agg = {
            r["jenis"]: r for r in _pr.filter(jenis__in=["depo", "wd"])
            .values("jenis").annotate(n=Count("id"), v=Sum("amount"))
        }
        _dp = _agg.get("depo") or {}
        _wd = _agg.get("wd") or {}
        dp_n, dp_v = _dp.get("n") or 0, float(_dp.get("v") or 0)
        wd_n, wd_v = _wd.get("n") or 0, float(_wd.get("v") or 0)
        panel_sum = {
            "dp": {"n": dp_n, "v": dp_v},
            "wd": {"n": wd_n, "v": wd_v},
            "total_n": dp_n + wd_n,
            "net": dp_v - wd_v,
        }
        # kartu "Metode Pembayaran": breakdown Bank Title dari queryset yang
        # SAMA (_pr) — satu query agregat kecil, total pasti klop panel_sum.
        metode = breakdown_metode(_pr)

    last_no = total_b if last else None
    last_sel = selisih(last) if last else 0
    pending = pending_settlement_count(active)
    um_d = {}
    if last is not None:
        um = (last.summary or {}).get("unmatched_money") or {}
        um_d = um.get("d") or {}

    comp = check_completeness(active)
    next_date = (last.recon_date + timedelta(days=1)) if last else today

    ctx = {
        "active_toko": active,
        "tx_total": tx.count(),
        "upload_total": uploads.count(),
        "run_total": runs.count(),
        "by_source": by_source,
        "uploads": uploads.select_related("source_type").order_by("-id")[:6],
        "runs": runs.select_related("batch").order_by("-id")[:6],
        "kal": kal,
        "tren": tren,
        "last": last, "last_no": last_no, "last_sel": last_sel,
        "panel_sum": panel_sum,
        "metode": metode,
        "pending": pending,
        "um_d": um_d,
        "comp": comp,
        "next_date": next_date,
    }
    return render(request, "web/dashboard.html", ctx)


# Upload folder/zip: hanya ekstensi yang punya parser; sisanya junk OS/temp.
_UPLOAD_EXTS = {".xlsx", ".xls", ".csv", ".pdf"}
_ZIP_MAX_FILES = 200
_ZIP_MAX_BYTES = 200 * 1024 * 1024
# Cap ukuran: per-file dan total satu request analyze — volume staging jangan
# bisa dipenuhi satu upload liar (export mutasi riil terbesar masih < 20MB).
_FILE_MAX_BYTES = 50 * 1024 * 1024
_REQ_MAX_BYTES = 300 * 1024 * 1024


def _is_junk_name(name):
    """File yang tak layak dianalisis: dotfile/.DS_Store, lock Office (~$),
    artefak __MACOSX, atau ekstensi tanpa parser. Berlaku untuk upload langsung,
    isi folder (webkitdirectory), dan isi zip."""
    path = str(name).replace("\\", "/")
    if "__MACOSX" in path:
        return True
    base = os.path.basename(path)
    if not base or base.startswith(".") or base.startswith("~$"):
        return True
    return os.path.splitext(base)[1].lower() not in _UPLOAD_EXTS


def _extract_zip(f):
    """Ekstrak arsip zip upload → (list[(nama, bytes)], n_dilewati, error|None).
    Guard: jumlah file & total ukuran terkompresi-buka (anti zip-bomb); zip
    berpassword/rusak → error berpesan jelas. xlsx TIDAK lewat sini (dicek
    berdasarkan ekstensi .zip, bukan magic PK — xlsx juga arsip zip)."""
    try:
        zf = zipfile.ZipFile(f)
    except zipfile.BadZipFile:
        return [], 0, "bukan file zip yang valid"
    infos = [i for i in zf.infolist() if not i.is_dir()]
    if len(infos) > _ZIP_MAX_FILES:
        return [], 0, f"terlalu banyak file di dalam zip (>{_ZIP_MAX_FILES})"
    if sum(i.file_size for i in infos) > _ZIP_MAX_BYTES:
        return [], 0, "isi zip terlalu besar (>200MB)"
    out, dilewati = [], 0
    for i in infos:
        if _is_junk_name(i.filename):
            dilewati += 1
            continue
        try:
            data = zf.read(i)
        except RuntimeError:
            return [], 0, "zip berpassword tidak didukung — ekstrak dulu lalu upload isinya"
        out.append((os.path.basename(i.filename.replace("\\", "/")), data))
    return out, dilewati, None


# File staging lebih tua dari ini = yatim (analyze tanpa commit) → disapu.
_STAGING_TTL = 24 * 3600


def _sweep_staging():
    """Bersihkan file staging yatim. Dipanggil tiap analyze — murah (satu
    listdir), dan berjalan tepat saat volume dipakai lagi."""
    import time

    try:
        _dirs, files = default_storage.listdir("staging")
    except FileNotFoundError:
        return
    batas = time.time() - _STAGING_TTL
    for nama in files:
        rel = f"staging/{nama}"
        try:
            if os.path.getmtime(default_storage.path(rel)) < batas:
                default_storage.delete(rel)
        except OSError:
            continue


def _analyze_file(name, fileobj):
    """Satu file → baris preview (simpan ke staging + deteksi parser).
    Dipakai upload langsung maupun hasil ekstrak zip."""
    saved = default_storage.save(f"staging/{name}", fileobj)
    needs_password = is_encrypted_xlsx(default_storage.path(saved))
    cands = detect_source(default_storage.path(saved), name)
    top = cands[0] if cands else None
    parser_key = top["parser_key"] if top else ""
    if not parser_key and needs_password:
        parser_key = "mandiri"
    return {
        "name": name,
        "staged": saved,
        "parser_key": parser_key,
        "confidence": round(top["confidence"] * 100) if top else 0,
        "needs_confirm": (top is None) or top["confidence"] < 0.8,
        "needs_password": needs_password,
        "flow": detect_flow(name),
    }


def _uploads_for(toko):
    """Riwayat upload toko (queryset penuh, terbaru dulu — view yang memaginasi),
    dianotasi `locked` (buktinya dipakai hasil rekon: direferensi MatchResult
    left/right ATAU dikonsumsi batch). Tombol Hapus per-baris dinonaktifkan;
    server (`_locking_batches`) tetap penjaga terakhir.

    left/right WAJIB dua Exists TERPISAH — bentuk lama `Q(left…) | Q(right…)`
    dalam SATU subquery memaksa Postgres OR lintas dua join (tak bisa pakai
    index, seq-scan MatchResult per baris upload): halaman Upload 10,8 dtk
    di prod; split Exists terukur 0,01 dtk dengan hasil identik."""
    ref_left = MatchResult.objects.filter(left__upload=OuterRef("pk"))
    ref_right = MatchResult.objects.filter(right__upload=OuterRef("pk"))
    consumed = Transaction.objects.filter(
        upload=OuterRef("pk"), consumed_by_batch__isnull=False
    )
    return (
        Upload.objects.filter(toko=toko)
        .select_related("source_type")
        .annotate(locked=ExpressionWrapper(
            Exists(ref_left) | Exists(ref_right) | Exists(consumed),
            output_field=BooleanField(),
        ))
        .order_by("-id")
    )


def _uploads_page(toko, request, q=""):
    """Halaman riwayat upload (20/halaman, pager seragam) — dulu terpotong 20
    terakhir sehingga file tanggal lama tak bisa dihapus dari UI.
    `q` (khusus admin, diisi view) menyaring nama file."""
    qs = _uploads_for(toko)
    if q:
        qs = qs.filter(original_name__icontains=q)
    return Paginator(qs, 20).get_page(request.GET.get("page"))


@login_required
def upload(request):
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    q = request.GET.get("q", "").strip() if is_admin(request.user) else ""
    if request.method == "POST" and request.POST.get("action") == "commit":
        staged = request.POST.getlist("staged")
        keys = request.POST.getlist("parser_key")
        flows = request.POST.getlist("flow")
        passwords = request.POST.getlist("password")
        provider = request.POST.get("provider", "")
        n_ok = n_err = 0
        for i, (path_rel, key, flow) in enumerate(zip(staged, keys, flows)):
            if not path_rel.startswith("staging/") or ".." in path_rel:
                n_err += 1
                continue
            if key not in PARSERS:
                n_err += 1
                continue
            try:
                ingest(
                    key, default_storage.path(path_rel), flow=flow,
                    user=request.user, toko=active, provider=provider,
                    password=(passwords[i] if i < len(passwords) else ""),
                )
                n_ok += 1
            except Exception as e:  # noqa: BLE001 - tampilkan error parse ke user
                messages.error(request, f"{path_rel}: {e}")
                n_err += 1
            finally:
                if default_storage.exists(path_rel):
                    default_storage.delete(path_rel)
        messages.success(request, f"{n_ok} file diproses, {n_err} gagal.")
        return redirect("upload")
    if request.method == "POST" and request.POST.get("action") == "analyze":
        _sweep_staging()
        uploaded = request.FILES.getlist("files")
        if sum(f.size for f in uploaded) > _REQ_MAX_BYTES:
            messages.error(request, "Total upload melebihi 300MB — pecah jadi beberapa kali.")
            uploaded = []
        preview, dilewati, diekstrak = [], 0, 0
        for f in uploaded:
            if f.size > _FILE_MAX_BYTES:
                messages.error(request, f"{f.name}: melebihi 50MB per file, dilewati.")
                dilewati += 1
                continue
            if f.name.lower().endswith(".zip"):
                isi, n_lewat, err = _extract_zip(f)
                if err:
                    messages.error(request, f"{f.name}: {err}")
                    continue
                dilewati += n_lewat
                for nama, data in isi:
                    preview.append(_analyze_file(nama, ContentFile(data)))
                    diekstrak += 1
                continue
            if _is_junk_name(f.name):
                dilewati += 1
                continue
            preview.append(_analyze_file(f.name, f))
        if dilewati or diekstrak:
            messages.info(
                request,
                f"{len(preview)} file dianalisa, {dilewati} dilewati"
                + (f", {diekstrak} diekstrak dari zip" if diekstrak else "")
                + ".",
            )
        return render(request, "web/upload.html", {
            "preview": preview, "parsers": sorted(PARSERS.keys()),
            "flows": ["", "dp", "wd"], "active_toko": active,
            "uploads": _uploads_page(active, request, q=q),
        })
    from reconciliation.engine import check_completeness

    return render(request, "web/upload.html", {
        "parsers": sorted(PARSERS.keys()), "active_toko": active,
        "uploads": _uploads_page(active, request, q=q),
        "comp": check_completeness(active),
        "q": q,
    })


@login_required
def transactions(request):
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    qs = (
        Transaction.objects.filter(toko=active)
        .select_related("source_type", "account", "upload", "upload__account")
    )
    src = request.GET.get("source", "")
    jenis = request.GET.get("jenis", "")
    q = request.GET.get("q", "").strip()
    bank = request.GET.get("bank", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    if src:
        qs = qs.filter(source_type__key=src)
    if jenis:
        qs = qs.filter(jenis=jenis)
    if q:
        cond = (
            Q(username__icontains=q)
            | Q(ticket_no__icontains=q)
            | Q(reference__icontains=q)
            | Q(counterparty__icontains=q)
        )
        # Angka (boleh berformat "50.000" / "50,000") → cari juga nominal persis.
        digits = q.replace(".", "").replace(",", "")
        if digits.isdigit():
            cond |= Q(amount=digits)
        qs = qs.filter(cond)
    carry = request.GET.get("carry") == "1"
    if carry:
        from reconciliation.engine import _carried_results

        qs = qs.filter(id__in=list(_carried_results(active).keys()))
    try:
        if date_from:
            qs = qs.filter(occurred_at__date__gte=date_cls.fromisoformat(date_from))
    except ValueError:
        date_from = ""
    try:
        if date_to:
            qs = qs.filter(occurred_at__date__lte=date_cls.fromisoformat(date_to))
    except ValueError:
        date_to = ""

    # Tombol filter per-bank: label diturunkan dari data upload toko ini
    # (account.provider / provider / nama file) — bukan daftar hardcode.
    bank_options = []
    if src in ("bank", "gateway"):
        ups = Upload.objects.filter(toko=active, source_type__key=src).select_related("account")
        label_by_upload = {
            u.id: specific_source_label(src, account=u.account, upload=u) for u in ups
        }
        fallback = src.capitalize()
        bank_options = sorted({lbl for lbl in label_by_upload.values() if lbl and lbl != fallback})
        if bank:
            qs = qs.filter(
                upload_id__in=[uid for uid, lbl in label_by_upload.items() if lbl == bank]
            )
    else:
        bank = ""

    qs, sort, sort_dir = _apply_sort(
        request, qs,
        allowed={
            "waktu": "occurred_at", "amount": "amount", "delta": "money_delta",
            "sumber": "source_type__key", "jenis": "jenis",
        },
        default_order=["-occurred_at", "id"],
        default_active=("waktu", "desc"),
    )

    if request.GET.get("export"):
        n = qs.count()
        if n > TX_EXPORT_LIMIT:
            messages.error(
                request,
                f"{n:,} baris terlalu banyak untuk diekspor — persempit filter dulu "
                f"(maks {TX_EXPORT_LIMIT:,}).",
            )
            # buang param export agar redirect tidak memicu export lagi (loop)
            redir = request.GET.copy()
            redir.pop("export", None)
            return redirect(f"{reverse('transactions')}?{redir.urlencode()}")
        return _export_transactions(qs, active)

    params = request.GET.copy()
    for k in ("sort", "dir", "page"):
        params.pop(k, None)
    qbase = params.urlencode()
    params_page = request.GET.copy()
    params_page.pop("page", None)
    qpage = params_page.urlencode()
    # Basis untuk tombol cepat Deposit/Withdraw: pertahankan filter lain,
    # tapi buang jenis/sort/dir/page agar tombol yang mengatur jenis.
    params_flt = request.GET.copy()
    for k in ("jenis", "sort", "dir", "page"):
        params_flt.pop(k, None)
    qflt = params_flt.urlencode()

    page = Paginator(qs, 40).get_page(request.GET.get("page"))

    # Ticket/Username/Nama Lengkap sisi uang: ambil dari pasangan panel/bracket
    # hasil rekonsiliasi — hanya untuk baris halaman ini (tanpa join tabel penuh).
    txs = list(page.object_list)
    page.object_list = txs
    money_ids = [t.id for t in txs if t.source_type.key in ("bank", "gateway")]
    best = {}
    if money_ids:
        results = (
            MatchResult.objects.filter(right_id__in=money_ids, left__isnull=False)
            .exclude(bucket=MatchResult.Bucket.TIDAK)
            .select_related("left")
        )
        for r in results:
            # cocok > skor tertinggi > run terbaru
            rank = (r.bucket == MatchResult.Bucket.COCOK, r.score or 0, r.run_id, r.id)
            if r.right_id not in best or rank > best[r.right_id][0]:
                best[r.right_id] = (rank, r.left)
    for t in txs:
        t.is_money = t.source_type.key in ("bank", "gateway")
        t.matched_panel = best.get(t.id, (None, None))[1]

    ctx = {
        "page": page,
        "sources": SourceType.objects.all(),
        "jenis_choices": Transaction.Jenis.choices,
        "src": src,
        "jenis": jenis,
        "q": q,
        "bank": bank,
        "bank_options": bank_options,
        "total": page.paginator.count,
        "date_from": date_from, "date_to": date_to,
        "sort": sort, "dir": sort_dir,
        "qbase": qbase, "qpage": qpage, "qflt": qflt,
        "carry": carry,
    }
    return render(request, "web/transactions.html", ctx)


def _export_transactions(qs, active):
    import io
    from datetime import datetime as _dt

    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font

    rows = list(qs)
    money_ids = [t.id for t in rows if t.source_type.key in ("bank", "gateway")]
    best = {}
    if money_ids:
        results = (
            MatchResult.objects.filter(right_id__in=money_ids, left__isnull=False)
            .exclude(bucket=MatchResult.Bucket.TIDAK)
            .select_related("left")
        )
        for r in results:
            rank = (r.bucket == MatchResult.Bucket.COCOK, r.score or 0, r.run_id, r.id)
            if r.right_id not in best or rank > best[r.right_id][0]:
                best[r.right_id] = (rank, r.left)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Transaksi")
    bold = Font(bold=True)

    def hcell(v):
        c = WriteOnlyCell(ws, value=v)
        c.font = bold
        return c

    ws.append([hcell(h) for h in [
        "Waktu", "Sumber", "Jenis", "Nominal", "Δ Uang", "Ticket",
        "Username", "Nama Lengkap", "Nama di Bank",
    ]])
    for t in rows:
        mp = best.get(t.id, (None, None))[1]
        is_money = t.source_type.key in ("bank", "gateway")
        ticket = t.ticket_no or (f"≈ {mp.ticket_no}" if mp and mp.ticket_no else "")
        username = t.username or (f"≈ {mp.username}" if mp and mp.username else "")
        if is_money:
            nama = f"≈ {mp.counterparty}" if mp and mp.counterparty else ""
        else:
            nama = t.counterparty or ""
        ws.append([
            t.occurred_at.strftime("%d/%m/%Y %H:%M") if t.occurred_at else "",
            t.source_label,
            t.get_jenis_display(),
            float(t.amount),
            float(t.money_delta),
            ticket, username, nama, t.counterparty or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fname = f"transaksi_{active.name}_{_dt.now():%Y%m%d-%H%M}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def reconcile(request):
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    if request.method == "POST":
        tol = get_object_or_404(ToleranceProfile, name=request.POST.get("tolerance", "Default"))
        # Checkbox inc_* per baris kelengkapan = sumber yang DIIKUTKAN. Tidak ada
        # → tidak dicentang → tidak dicocokkan & tidak dikonsumsi.
        include = {
            "panel_dp": "inc_panel_dp" in request.POST,
            "panel_wd": "inc_panel_wd" in request.POST,
            "bracket": "inc_bracket" in request.POST,
            "bank": "inc_bank" in request.POST,
            "gateway": "inc_gateway" in request.POST,
        }
        # Auto-split: satu batch per tanggal-panel. Panel jadi jangkar; run ditolak
        # bila ada uang/bracket tanpa panel penutup dalam window (verify_panel_anchor).
        res = run_batches_auto(
            active, tol,
            request.POST.get("date_from") or None,
            request.POST.get("date_to") or None,
            user=request.user, include=include,
        )
        if not res["ok"]:
            rows = format_html_join(
                "", "<br>&bull; {} — {} ({} baris)",
                ((v["date"].strftime("%d/%m/%Y"), v["source"], v["n"]) for v in res["violations"]),
            )
            messages.error(request, format_html(
                "Rekonsiliasi ditolak: ada tanggal ber-uang/bracket tanpa panel penutup. "
                "Upload panel tanggal terkait dulu, lalu jalankan lagi:{}", rows,
            ))
            return redirect("reconcile")
        for er in res["errors"]:
            messages.error(request, f"{er['date'].strftime('%d/%m/%Y')}: {er['message']}")
        batches, skipped = res["batches"], res["skipped_existing"]
        for b in batches:
            no = ReconBatch.objects.filter(toko=active, id__lte=b.id).count()
            catat(request.user, "reconcile", f"Batch #{no}", toko=active, batch_pk=b.pk)
        if len(batches) == 1 and not skipped:
            no = ReconBatch.objects.filter(toko=active).count()
            messages.success(request, f"Rekonsiliasi selesai (Batch #{no}).")
            return redirect("batch_detail", pk=batches[0].pk)
        if batches:
            rentang = (
                f"{batches[0].recon_date.strftime('%d/%m')}–"
                f"{batches[-1].recon_date.strftime('%d/%m/%Y')}"
                if len(batches) > 1 else batches[0].recon_date.strftime("%d/%m/%Y")
            )
            msg = f"{len(batches)} batch dibuat ({rentang})."
            if skipped:
                msg += f" {len(skipped)} tanggal dilewati (sudah ada batch)."
            messages.success(request, msg)
        elif skipped:
            messages.info(
                request,
                f"Semua {len(skipped)} tanggal sudah punya batch — tak ada yang baru dibuat.",
            )
        else:
            messages.info(request, "Tidak ada tanggal panel untuk diproses. Upload panel dulu.")
        return redirect("reconcile")

    df = request.GET.get("date_from") or None
    dt = request.GET.get("date_to") or None
    bank = request.GET.get("bank", "")
    if bank not in ("bank", "gateway"):
        bank = ""  # nilai tak dikenal → perlakukan sebagai "semua sumber"
    # Nomor dihitung dari SEMUA batch toko dulu (posisi asli), BARU difilter —
    # supaya nomor batch tidak berubah saat filter sumber uang aktif.
    all_batches = list(ReconBatch.objects.filter(toko=active).order_by("-id"))
    total = len(all_batches)
    for i, b in enumerate(all_batches):
        b.no = total - i
    if bank:
        all_batches = [b for b in all_batches if (b.completeness or {}).get(bank)]
    batches = all_batches[:20]
    comp = check_completeness(active, df, dt)
    comp_keys = ["panel_dp", "panel_wd", "bracket", "bank", "gateway"]
    comp_ready = sum(1 for k in comp_keys if comp.get(k))
    # Preview auto-split: tanggal-panel yang akan diproses (tiap tanggal = satu batch).
    panel_dates = _panel_dates(active, df, dt, None)
    ctx = {
        "active_toko": active,
        "completeness": comp,
        "comp_ready": comp_ready,
        "comp_total": len(comp_keys),
        "comp_pct": round(100 * comp_ready / len(comp_keys)),
        "tolerances": ToleranceProfile.objects.all(),
        "batches": batches,
        "bank": bank,
        "date_from": df or "", "date_to": dt or "",
        "panel_dates": panel_dates,
        "panel_dates_count": len(panel_dates),
        "pending_settlement": pending_settlement_count(active),
    }
    return render(request, "web/reconcile.html", ctx)


@login_required
def batch_detail(request, pk):
    batch = get_object_or_404(ReconBatch, pk=pk, toko__in=tokos_for(request.user))
    batch_no = ReconBatch.objects.filter(toko=batch.toko, id__lte=batch.id).count()
    # Settle terlambat dua arah — dari queryset LIVE (bukan summary JSON) supaya
    # otomatis kosong bila batch pasangannya sudah dihapus.
    resolved_here = list(
        MatchResult.objects.filter(resolved_by_batch=batch)
        .select_related("left", "right", "run__batch")
    )
    for r in resolved_here:  # nomor batch asal (konvensi nomor per-toko)
        r.home_no = ReconBatch.objects.filter(
            toko=batch.toko, id__lte=r.run.batch_id
        ).count()
    settled_elsewhere = list(
        MatchResult.objects.filter(run__batch=batch, resolved_by_batch__isnull=False)
        .select_related("resolved_by_batch", "left", "right")
    )

    # Ringkasan per sumber uang (dihitung on-the-fly; batch lama otomatis kebagian).
    from django.db.models import Exists, OuterRef

    from transactions.models import specific_source_label

    paired_q = MatchResult.objects.filter(left__isnull=False, right_id=OuterRef("id"))
    money_rows = (
        Transaction.objects.filter(
            consumed_by_batch=batch, source_type__key__in=["bank", "gateway"]
        )
        .exclude(jenis="admin")
        .annotate(berpasangan=Exists(paired_q))
        .select_related("source_type", "upload", "upload__account")
    )
    agg = {}
    for t in money_rows:
        label = specific_source_label(
            t.source_type.key,
            account=t.upload.account if t.upload else None,
            upload=t.upload,
        )
        row = agg.setdefault(
            label, {"label": label, "n": 0, "dp": 0.0, "wd": 0.0, "paired": 0, "unpaired": 0}
        )
        row["n"] += 1
        md = float(t.money_delta)
        if md > 0:
            row["dp"] += md
        elif md < 0:
            row["wd"] += -md
        if t.berpasangan:
            row["paired"] += 1
        else:
            row["unpaired"] += 1
    per_bank = sorted(agg.values(), key=lambda r: r["label"])

    # Deteksi cangkang (F1-B): summary mengklaim hasil tapi MatchResult sudah tak
    # ada (bukti terhapus → cascade). Batch begini tak bisa diverifikasi.
    bkt = (batch.summary or {}).get("buckets", {})
    claimed = sum(v for v in bkt.values() if isinstance(v, (int, float)))
    is_hollow = claimed > 0 and not MatchResult.objects.filter(run__batch=batch).exists()
    riwayat = list(
        AuditLog.objects.filter(detail__batch_pk=batch.pk).select_related("user")[:20]
    )

    return render(request, "web/batch_detail.html", {
        "batch": batch, "batch_no": batch_no, "s": batch.summary or {}, "runs": batch.runs.all(),
        "resolved_here": resolved_here, "settled_elsewhere": settled_elsewhere,
        "per_bank": per_bank, "is_hollow": is_hollow, "riwayat": riwayat,
    })


KATEGORI_UANG = {
    "a": ("Histori", "di luar periode rekonsiliasi"),
    "b": ("Ticket asing", "ticket gateway tak dikenal panel"),
    "c": ("Internal", "pindah dana antar rekening operator"),
    "d": ("Periksa", "dalam periode tanpa catatan panel"),
}


@login_required
def batch_uang(request, pk):
    """Uang tanpa pasangan milik satu batch — daftar live berkategori a/b/c/d.
    Baris b/d juga punya MatchResult no_panel (bisa ditinjau di halaman run);
    halaman ini adalah ikhtisar + filter + export."""
    from django.db.models import Exists, OuterRef

    from reconciliation.engine import _operator_names, classify_unmatched_money

    batch = get_object_or_404(ReconBatch, pk=pk, toko__in=tokos_for(request.user))
    batch_no = ReconBatch.objects.filter(toko=batch.toko, id__lte=batch.id).count()
    paired = MatchResult.objects.filter(left__isnull=False, right_id=OuterRef("id"))
    rows = list(
        Transaction.objects.filter(
            consumed_by_batch=batch, source_type__key__in=["bank", "gateway"]
        )
        .exclude(jenis="admin")
        .annotate(berpasangan=Exists(paired))
        .filter(berpasangan=False)
        .select_related("source_type", "upload")
        .order_by("occurred_at", "id")
    )
    recon_date = batch.recon_date
    window = batch.tolerance.date_window_days
    if recon_date:
        panel_tickets = set(
            Transaction.objects.filter(toko=batch.toko, source_type__key="panel")
            .exclude(ticket_no="").values_list("ticket_no", flat=True)
        )
        ops = _operator_names(batch.toko)
        for t in rows:
            t.kategori = classify_unmatched_money(t, recon_date, window, panel_tickets, ops)
    else:  # batch lama tanpa tanggal harian — tak bisa diklasifikasi
        for t in rows:
            t.kategori = "d"
    stats = {k: {"n": 0, "amt": 0.0} for k in KATEGORI_UANG}
    for t in rows:
        stats[t.kategori]["n"] += 1
        stats[t.kategori]["amt"] += abs(float(t.money_delta))
    kat = request.GET.get("k", "")
    if kat in KATEGORI_UANG:
        rows = [t for t in rows if t.kategori == kat]

    # Sort in-memory (rows sudah list, kategori dihitung pasca-query).
    from datetime import datetime as _dt

    sort = request.GET.get("sort", "")
    sort_dir = request.GET.get("dir", "")
    if sort == "nominal":
        rows.sort(key=lambda t: abs(float(t.money_delta)), reverse=(sort_dir == "desc"))
    elif sort == "tanggal":
        rows.sort(key=lambda t: (t.occurred_at or _dt.min), reverse=(sort_dir == "desc"))
    else:
        sort, sort_dir = "tanggal", "asc"  # default: tanggal naik
        rows.sort(key=lambda t: (t.occurred_at or _dt.min))

    if request.GET.get("export"):
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Uang tanpa pasangan"
        ws.append(["Kategori", "Tanggal", "Sumber", "File/Rekening", "Ticket",
                   "Username", "Pengirim/Penerima", "Nominal"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for t in rows:
            ws.append([
                t.kategori.upper(),
                t.occurred_at.strftime("%d/%m/%Y %H:%M") if t.occurred_at else "",
                t.source_type.key,
                t.upload.original_name if t.upload else "",
                t.ticket_no, t.username, t.counterparty,
                float(t.money_delta),
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="uang_tanpa_pasangan_batch{batch.pk}.xlsx"'
        )
        return resp

    page = Paginator(rows, 40).get_page(request.GET.get("page"))
    kartu = [
        {"key": k, "label": KATEGORI_UANG[k][0], "desc": KATEGORI_UANG[k][1],
         "n": stats[k]["n"], "amt": stats[k]["amt"]}
        for k in KATEGORI_UANG
    ]
    return render(request, "web/batch_uang.html", {
        "batch": batch, "batch_no": batch_no, "page": page,
        "kartu": kartu, "kat": kat, "sort": sort, "dir": sort_dir,
    })


@login_required
def run_detail(request, pk):
    run = get_object_or_404(MatchRun, pk=pk, batch__toko__in=tokos_for(request.user))
    base = MatchResult.objects.filter(run=run).select_related(
        "left", "right", "right__source_type", "right__upload", "right__account"
    )

    # Kartu status: 'Tidak Cocok' (masih ada baris kredit / no_money) dipisah dari
    # 'Tidak Ada di Panel' (orphan uang tanpa kredit / no_panel) — dihitung live.
    TIDAK = MatchResult.Bucket.TIDAK
    n_tidak_cocok = base.filter(bucket=TIDAK, left__isnull=False).count()
    n_tidak_ada_panel = base.filter(bucket=TIDAK, left__isnull=True).count()

    qs = base
    bucket = request.GET.get("bucket", "")
    if bucket == "tidak_cocok":
        qs = qs.filter(bucket=TIDAK, left__isnull=False)
    elif bucket == "tidak_ada_panel":
        qs = qs.filter(bucket=TIDAK, left__isnull=True)
    elif bucket:
        qs = qs.filter(bucket=bucket)

    # Filter arus Deposit/Withdraw: pakai jenis sisi kiri; hasil tanpa kiri
    # (mis. no_panel) dinilai dari sisi uangnya.
    flow = request.GET.get("flow", "")
    if flow not in ("depo", "wd"):
        flow = ""
    if flow:
        qs = qs.filter(Q(left__jenis=flow) | Q(left__isnull=True, right__jenis=flow))

    # Chip filter per alasan — dihitung DALAM bucket+flow terpilih supaya angkanya jujur.
    reasons = list(
        qs.values("reason_code").annotate(n=Count("id")).order_by("-n")
    )
    reason = request.GET.get("reason", "")
    if reason:
        qs = qs.filter(reason_code=reason)

    # Chip filter bank pemain (dalam bucket+flow+alasan). Kosong/orphan dikecualikan.
    banks = [
        {"code": r["left__player_bank"], "n": r["n"]}
        for r in qs.filter(left__player_bank__gt="")
        .values("left__player_bank").annotate(n=Count("id")).order_by("-n")
    ]
    bank = request.GET.get("bank", "")
    if bank:
        qs = qs.filter(left__player_bank=bank)

    # Chip filter bank title / tujuan (dalam bucket+flow+alasan+bank).
    btitles = [
        {"code": r["left__bank_title"], "n": r["n"]}
        for r in qs.filter(left__bank_title__gt="")
        .values("left__bank_title").annotate(n=Count("id")).order_by("-n")
    ]
    btitle = request.GET.get("btitle", "")
    if btitle:
        qs = qs.filter(left__bank_title=btitle)

    # Ringkasan total pada set terfilter penuh (sebelum paginasi).
    totals = qs.aggregate(
        kredit=Sum("left__amount"), saldo=Sum("right__amount"), n=Count("id")
    )

    qs, sort, sort_dir = _apply_sort(
        request, qs,
        allowed={"amount": "left__amount", "skor": "score", "waktu": "left__occurred_at"},
        default_order=["bucket", "-score", "id"],
    )
    page = Paginator(_annot_alasan_review(qs), 40).get_page(request.GET.get("page"))
    left_label, right_label = REL_LABELS.get(run.relation, ("Kiri", "Kanan"))
    left_amt_label, right_amt_label = REL_AMOUNT_LABELS.get(
        run.relation, ("Kredit/Koin", "Saldo Bank")
    )
    orphan_label = f"Tidak Ada di {left_label}"
    # Nomor batch per-toko (posisi urut, bukan pk global) — konsisten dgn batch_detail.
    batch = run.batch
    batch_no = (
        ReconBatch.objects.filter(toko=batch.toko, id__lte=batch.id).count() if batch else None
    )
    # Cangkang (F1-B): summary run mengklaim hasil tapi MatchResult sudah tak ada.
    rs = run.summary or {}
    claimed = sum(v for v in rs.values() if isinstance(v, (int, float)))
    is_hollow = claimed > 0 and not MatchResult.objects.filter(run=run).exists()
    ctx = {
        "run": run, "page": page, "bucket": bucket, "bucket_meta": BUCKET_META,
        "left_label": left_label, "right_label": right_label,
        "left_amt_label": left_amt_label, "right_amt_label": right_amt_label,
        "orphan_label": orphan_label,
        "n_tidak_cocok": n_tidak_cocok, "n_tidak_ada_panel": n_tidak_ada_panel,
        # Tab "Tidak Ada di Panel": sisi kiri (panel) kosong semua → sembunyikan
        # kolomnya agar kolom Mutasi Bank lega & nama rekening terbaca.
        "hide_left": bucket == "tidak_ada_panel",
        "batch": batch, "batch_no": batch_no,
        "reasons": reasons, "reason": reason, "flow": flow,
        "banks": banks, "bank": bank, "btitles": btitles, "btitle": btitle,
        "totals": totals,
        "sort": sort, "dir": sort_dir, "is_hollow": is_hollow,
        "pilihan_alasan": FRKoreksi.ALASAN_KOREKSI,
    }
    return render(request, "web/run_detail.html", ctx)


def _parse_date(s):
    """'YYYY-MM-DD' → date; string kosong/invalid → None (filter diabaikan)."""
    try:
        return date_cls.fromisoformat(s) if s else None
    except ValueError:
        return None


@login_required
def review_queue(request):
    """Area Pengecekan: hasil yang perlu dicek toko aktif, lintas batch/run —
    tab perlu_tinjau (default) / tidak_cocok / tidak_ada_panel.
    Filter (pola sama dgn run_detail): DP/WD, rentang tanggal transaksi,
    bank pemain, bank title — chip dihitung DALAM filter terpilih."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    TIDAK = MatchResult.Bucket.TIDAK
    bucket = request.GET.get("bucket", "perlu_tinjau")
    if bucket not in ("perlu_tinjau", "tidak_cocok", "tidak_ada_panel"):
        bucket = "perlu_tinjau"  # param ngawur -> default (back-compat URL lama)
    base = MatchResult.objects.filter(run__batch__toko=active).select_related(
        "left", "right", "right__source_type", "right__upload", "run", "run__batch"
    )
    flow = request.GET.get("flow", "")
    if flow not in ("depo", "wd"):
        flow = ""
    if flow:
        base = base.filter(Q(left__jenis=flow) | Q(left__isnull=True, right__jenis=flow))

    date_from = _parse_date(request.GET.get("from", ""))
    date_to = _parse_date(request.GET.get("to", ""))
    # Tanggal: sisi kredit; baris orphan (tanpa kiri) dinilai dari sisi uangnya.
    if date_from:
        base = base.filter(
            Q(left__occurred_at__date__gte=date_from)
            | Q(left__isnull=True, right__occurred_at__date__gte=date_from)
        )
    if date_to:
        base = base.filter(
            Q(left__occurred_at__date__lte=date_to)
            | Q(left__isnull=True, right__occurred_at__date__lte=date_to)
        )

    # Hitungan tab DALAM flow+tanggal terpilih (angka jujur, pola run_detail).
    tab_counts = {
        "perlu_tinjau": base.filter(bucket=MatchResult.Bucket.TINJAU).count(),
        "tidak_cocok": base.filter(bucket=TIDAK, left__isnull=False).count(),
        "tidak_ada_panel": base.filter(bucket=TIDAK, left__isnull=True).count(),
    }
    if bucket == "tidak_cocok":
        qs = base.filter(bucket=TIDAK, left__isnull=False)
    elif bucket == "tidak_ada_panel":
        # skor tak bermakna utk orphan -> urut waktu sisi uang
        qs = base.filter(bucket=TIDAK, left__isnull=True).order_by(
            "-run__batch__recon_date", "right__occurred_at", "id"
        )
    else:
        qs = base.filter(bucket=MatchResult.Bucket.TINJAU)
    if bucket != "tidak_ada_panel":
        qs = qs.order_by("-run__batch__recon_date", "-score", "id")

    banks = [
        {"code": r["left__player_bank"], "n": r["n"]}
        for r in qs.filter(left__player_bank__gt="")
        .values("left__player_bank").annotate(n=Count("id")).order_by("-n")
    ]
    bank = request.GET.get("bank", "")
    if bank:
        qs = qs.filter(left__player_bank=bank)

    btitles = [
        {"code": r["left__bank_title"], "n": r["n"]}
        for r in qs.filter(left__bank_title__gt="")
        .values("left__bank_title").annotate(n=Count("id")).order_by("-n")
    ]
    btitle = request.GET.get("btitle", "")
    if btitle:
        qs = qs.filter(left__bank_title=btitle)

    # Ringkasan total pada set terfilter penuh (sebelum paginasi) — pola run_detail.
    totals = qs.aggregate(
        kredit=Sum("left__amount"), saldo=Sum("right__amount"), n=Count("id")
    )

    page = Paginator(_annot_alasan_review(qs), 40).get_page(request.GET.get("page"))
    # nomor batch per-toko untuk tiap hasil di halaman ini
    for r in page.object_list:
        b = r.run.batch
        r.home_no = (
            ReconBatch.objects.filter(toko=active, id__lte=b.id).count() if b else None
        )
    return render(request, "web/review_queue.html", {
        "page": page, "active_toko": active,
        "bucket": bucket, "tab_counts": tab_counts, "totals": totals,
        "hide_left": bucket == "tidak_ada_panel",
        "flow": flow, "bank": bank, "btitle": btitle,
        "banks": banks, "btitles": btitles,
        "date_from": request.GET.get("from", "") if date_from else "",
        "date_to": request.GET.get("to", "") if date_to else "",
        "pilihan_alasan": FRKoreksi.ALASAN_KOREKSI,
    })


# No. HP di keterangan mutasi e-wallet (GoPay/DANA): '085767555197' atau
# '82279003062' (tanpa 0). \b menjaga kode alfanumerik (WS95011...) tak ikut.
PHONE_RE = re.compile(r"\b0?8\d{8,12}\b")


def _resolve_wallet_names(rows, toko):
    """Baris mutasi tanpa nama (e-wallet, hanya HP): tempelkan r.phone + r.player_name.

    Nama dicari di panel toko yang sama via segmen HP di raw['Player Bank']
    ('KODE|NAMA|ACCT' — lihat parse_bank_triplet). Per halaman saja (<=40 baris,
    <=40 query) — bukan jalur matching, murni tampilan.
    """
    for r in rows:
        if r.counterparty or r.source_type.key not in ("bank", "gateway"):
            continue
        m = PHONE_RE.search(r.description or "")
        if not m:
            continue
        r.phone = m.group(0)
        suffix = r.phone.lstrip("0")
        cand = (
            Transaction.objects.filter(
                toko=toko, source_type__key="panel",
                **{"raw__Player Bank__icontains": suffix},
            )
            .order_by("-occurred_at")
            .first()
        )
        if cand:
            r.player_name = cand.counterparty or cand.username


@login_required
def bank_mutations(request):
    """Sub-menu Mutasi Bank: baris mutasi bank + gateway (QRIS) apa adanya,
    urut PERSIS file asli — grup per upload terbaru, di dalam file mengikuti
    urutan parse (bulk_create mempertahankan urutan -> id menaik)."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    money_keys = ("bank", "gateway")
    qs = (
        Transaction.objects.filter(toko=active, source_type__key__in=money_keys)
        .select_related("source_type", "upload", "account", "upload__account")
        .order_by("-upload_id", "id")
    )
    src = request.GET.get("source", "")
    if src not in money_keys:
        src = ""
    if src:
        qs = qs.filter(source_type__key=src)

    flow = request.GET.get("flow", "")
    if flow not in ("depo", "wd"):
        flow = ""
    if flow:
        qs = qs.filter(jenis=flow)

    date_from = _parse_date(request.GET.get("from", ""))
    date_to = _parse_date(request.GET.get("to", ""))
    if date_from:
        qs = qs.filter(occurred_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(occurred_at__date__lte=date_to)

    # Dropdown per-file: upload sumber uang toko aktif, IKUT tombol sumber
    # (Bank → hanya file bank; Gateway QRIS → hanya file gateway).
    upload_qs = Upload.objects.filter(toko=active, source_type__key__in=money_keys)
    if src:
        upload_qs = upload_qs.filter(source_type__key=src)
    uploads = list(upload_qs.order_by("-id"))
    # Rentang isi NYATA per file (satu query agregat, bukan per upload).
    # File ekspor bank sering rolling/tumpang-tindih: baris duplikat di-skip
    # dedup dan tercatat di upload TERDAHULU, jadi file baru bisa berisi
    # hanya "ekor"-nya — tanpa rentang ini user mengira mutasinya kepotong
    # (kasus nyata LBS 10/07: 697 baris file -> 47 baru, sisanya di file 08-09/07).
    cover = {
        r["upload_id"]: (r["lo"], r["hi"])
        for r in Transaction.objects.filter(upload__in=uploads)
        .values("upload_id")
        .annotate(lo=Min("occurred_at"), hi=Max("occurred_at"))
    }
    # Baris duplikat ter-link (isi file yang tercatat di upload terdahulu):
    # ikut dihitung agar jumlah baris + rentang dropdown = ISI FILE utuh.
    through = Upload.duplicate_transactions.through
    linked = {
        r["upload_id"]: (r["n"], r["lo"], r["hi"])
        for r in through.objects.filter(upload__in=uploads)
        .values("upload_id")
        .annotate(
            n=Count("transaction_id"),
            lo=Min("transaction__occurred_at"),
            hi=Max("transaction__occurred_at"),
        )
    }
    for u in uploads:
        lo, hi = cover.get(u.id, (None, None))
        n_link, dlo, dhi = linked.get(u.id, (0, None, None))
        u.n_dup_linked = n_link
        u.n_rows_file = u.rows_parsed + n_link
        u.cover_lo = min((d for d in (lo, dlo) if d), default=None)
        u.cover_hi = max((d for d in (hi, dhi) if d), default=None)
    upload_id = request.GET.get("upload", "")
    sel_upload = None
    if upload_id.isdigit():
        # cari di daftar ter-scope src → ganti sumber otomatis mereset pilihan file
        sel_upload = next((u for u in uploads if u.id == int(upload_id)), None)
        if sel_upload:  # id upload toko lain / sumber lain diabaikan (RBAC + konsistensi)
            dup_ids = list(sel_upload.duplicate_transactions.values_list("id", flat=True))
            if dup_ids:
                # ISI FILE UTUH: baris milik file + baris yang di-skip dedup
                # (tercatat di upload terdahulu). Urut waktu — posisi asli baris
                # duplikat di file tak tersimpan, dan ekspor bank kronologis.
                # id__in list terbatas (≤ isi file), BUKAN OR lintas-join.
                qs = qs.filter(Q(upload=sel_upload) | Q(id__in=dup_ids)).order_by(
                    "occurred_at", "id"
                )
            else:
                qs = qs.filter(upload=sel_upload)

    page = Paginator(qs, 40).get_page(request.GET.get("page"))
    _resolve_wallet_names(page.object_list, active)
    return render(request, "web/mutasi_bank.html", {
        "page": page, "active_toko": active,
        "src": src, "flow": flow,
        "uploads": uploads, "sel_upload": sel_upload,
        "date_from": request.GET.get("from", "") if date_from else "",
        "date_to": request.GET.get("to", "") if date_to else "",
    })


@login_required
def bracket_breakdown(request):
    """Sub-menu Breakdown Bracket: isi FR harian per FR Account — kartu
    "Pergerakan per Bank" (rekap DP/WD/Net/Trx/Saldo) + "Control Bracket
    Transaction" (pivot kategori asli FR + Selisih Kontrol). Lihat
    docs/superpowers/specs/2026-07-12-breakdown-bracket-design.md."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    latest = Transaction.objects.filter(
        toko=active, source_type__key="bracket"
    ).aggregate(m=Max("posted_date"))["m"]
    # back-compat: ?date= lama = ?dari=&sampai= (rentang 1 hari)
    lama = _parse_date(request.GET.get("date", ""))
    sampai = _parse_date(request.GET.get("sampai", "")) or lama or latest or date_cls.today()
    dari = _parse_date(request.GET.get("dari", "")) or lama or sampai
    if dari > sampai:
        dari, sampai = sampai, dari
    span = (sampai - dari).days + 1  # geser prev/next seluruh lebar jendela
    koreksi_on = dari == sampai
    data = hitung_bracket_breakdown(active, dari, sampai)
    return render(request, "web/breakdown_bracket.html", {
        "data": data, "dari": dari, "sampai": sampai, "tanggal": dari,
        "latest": latest, "koreksi_on": koreksi_on,
        "prev_dari": dari - timedelta(days=span),
        "prev_sampai": sampai - timedelta(days=span),
        "next_dari": dari + timedelta(days=span),
        "next_sampai": sampai + timedelta(days=span),
    })


_FR_KOLOM_SALDO = {"saldo_awal": "Saldo Awal", "saldo_akhir": "Saldo Akhir"}


def _fr_label_kolom(kolom):
    if kolom in _FR_KOLOM_SALDO:
        return _FR_KOLOM_SALDO[kolom]
    return dict(KATEGORI_KANONIK).get(kolom, kolom.title())


def _fr_asli(toko, tanggal, account, kolom):
    """(nilai asli sel, kolom sah?) dari agregasi MENTAH (tanpa koreksi).

    Kolom sah = saldo_awal/saldo_akhir, slug kanonik, atau kategori yang
    benar-benar muncul di data hari itu — menolak POST rakitan yang mau
    menyuntik kolom fiktif ke tabel kontrol.
    """
    data = hitung_bracket_breakdown(toko, tanggal, dengan_koreksi=False)
    sah = (kolom in _FR_KOLOM_SALDO or kolom in dict(KATEGORI_KANONIK)
           or kolom in {s for s, _ in data["kolom"]})
    asli = None
    for acc in data["accounts"]:
        if acc["account"] == account:
            if kolom in _FR_KOLOM_SALDO:
                asli = acc[kolom]
            else:
                asli = acc["kategori"].get(kolom)
            break
    return asli, sah


def _fr_params(request, src):
    tanggal = _parse_date(src.get("date", ""))
    account = (src.get("account") or "").strip()[:255]
    kolom = (src.get("kolom") or "").strip().lower()[:64]
    if not tanggal or not account or not kolom:
        return None
    return tanggal, account, kolom


@login_required
def fr_koreksi_form(request):
    """Popup kecil koreksi satu sel Control Bracket (GET, HTMX)."""
    active = _active_toko(request)
    params = _fr_params(request, request.GET)
    if active is None or params is None:
        return HttpResponseBadRequest("parameter kurang")
    tanggal, account, kolom = params
    asli, sah = _fr_asli(active, tanggal, account, kolom)
    if not sah:
        return HttpResponseBadRequest("kolom tidak dikenal")
    koreksi = FRKoreksi.objects.filter(
        toko=active, tanggal=tanggal, account=account, kolom=kolom).first()
    return render(request, "web/_fr_koreksi_form.html", {
        "tanggal": tanggal, "account": account, "kolom": kolom,
        "label": _fr_label_kolom(kolom),
        "asli": asli,
        "koreksi": koreksi, "pilihan_alasan": FRKoreksi.ALASAN_KOREKSI,
    })


@login_required
def fr_koreksi_simpan(request):
    """Simpan/hapus koreksi sel FR lalu render ulang tabel kontrol (POST, HTMX)."""
    if request.method != "POST":
        return HttpResponseBadRequest("POST saja")
    active = _active_toko(request)
    params = _fr_params(request, request.POST)
    if active is None or params is None:
        return HttpResponseBadRequest("parameter kurang")
    tanggal, account, kolom = params
    asli, sah = _fr_asli(active, tanggal, account, kolom)
    if not sah:
        return HttpResponseBadRequest("kolom tidak dikenal")

    if request.POST.get("hapus"):
        FRKoreksi.objects.filter(
            toko=active, tanggal=tanggal, account=account, kolom=kolom).delete()
        catat(request.user, "fr_koreksi_hapus", f"{account} [{kolom}]", toko=active,
              tanggal=str(tanggal), account=account, kolom=kolom,
              nilai_asli=str(asli) if asli is not None else "")
    else:
        mentah = (request.POST.get("nilai") or "").strip().replace(" ", "")
        try:
            # input polos tanpa pemisah ribuan; koma desimal diterima
            nilai = Decimal(mentah.replace(".", "").replace(",", "."))
        except InvalidOperation:
            return HttpResponseBadRequest("nilai tidak valid")
        if not nilai.is_finite():
            return HttpResponseBadRequest("nilai tidak valid")
        try:
            # bentuk eksponen raksasa (1e30) lolos is_finite tapi meledak di
            # quantize (melebihi presisi konteks) — tolak, jangan 500
            nilai = nilai.quantize(Decimal("0.01"))
        except InvalidOperation:
            return HttpResponseBadRequest("nilai tidak valid")
        if abs(nilai) > Decimal("9999999999999999.99"):
            return HttpResponseBadRequest("nilai tidak valid")
        alasan = request.POST.get("alasan") or ""
        if alasan and alasan not in dict(FRKoreksi.ALASAN_KOREKSI):
            return HttpResponseBadRequest("alasan tidak dikenal")
        FRKoreksi.objects.update_or_create(
            toko=active, tanggal=tanggal, account=account, kolom=kolom,
            defaults={"nilai": nilai, "alasan": alasan,
                      "catatan": (request.POST.get("catatan") or "").strip(),
                      "dibuat_oleh": request.user})
        catat(request.user, "fr_koreksi", f"{account} [{kolom}]", toko=active,
              tanggal=str(tanggal), account=account, kolom=kolom,
              nilai_asli=str(asli) if asli is not None else "",
              nilai_baru=str(nilai), alasan=alasan)

    data = hitung_bracket_breakdown(active, tanggal)
    html = render_to_string("web/_fr_control_table.html",
                            {"data": data, "tanggal": tanggal, "koreksi_on": True},
                            request=request)
    html += '<div id="koreksiPop" hx-swap-oob="innerHTML"></div>'
    gerak = render_to_string("web/_fr_gerak_table.html",
                             {"data": data, "tanggal": tanggal}, request=request)
    gerak = gerak.replace('id="fr-gerak"', 'id="fr-gerak" hx-swap-oob="outerHTML"', 1)
    html += gerak
    return HttpResponse(html)


@login_required
def hutang_piutang(request):
    """Daftar hutang/piutang FR lintas tanggal (otomatis dari data bracket)."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    sampai = _parse_date(request.GET.get("sampai", "")) or date_cls.today()
    dari = _parse_date(request.GET.get("dari", "")) or sampai - timedelta(days=30)
    data = hitung_hutang_piutang(active, dari=dari, sampai=sampai)
    page = Paginator(data["rows"], 40).get_page(request.GET.get("page"))
    return render(request, "web/hutang_piutang.html", {
        "page": page, "data": data, "dari": dari, "sampai": sampai,
    })


@login_required
def rincian_biaya(request):
    """Rekap biaya admin bank per kanal (E-wallet/BI Fast/Transfer online)."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    sampai = _parse_date(request.GET.get("sampai", "")) or date_cls.today()
    dari = _parse_date(request.GET.get("dari", "")) or sampai - timedelta(days=30)
    data = hitung_rincian_biaya(active, dari=dari, sampai=sampai)
    page = Paginator(data["rows"], 40).get_page(request.GET.get("page"))
    return render(request, "web/biaya_admin.html", {
        "page": page, "data": data, "dari": dari, "sampai": sampai,
    })


@login_required
def bonus_recon(request):
    """Rekonsiliasi Bonus panel<->bracket — tab panel_only (default)/bracket_only/cocok."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    sampai = _parse_date(request.GET.get("sampai", "")) or date_cls.today()
    dari = _parse_date(request.GET.get("dari", "")) or sampai - timedelta(days=30)
    kategori = (request.GET.get("kategori") or "").strip()
    data = hitung_rekonsiliasi_bonus(active, dari=dari, sampai=sampai,
                                     kategori=kategori or None)
    tab = request.GET.get("tab") or "panel"
    rows_by_tab = {
        "panel": data["panel_only"],
        "bracket": data["bracket_only"],
        "cocok": data["cocok"],
    }
    if tab not in rows_by_tab:
        tab = "panel"
    page = Paginator(rows_by_tab[tab], 40).get_page(request.GET.get("page"))
    return render(request, "web/bonus_recon.html", {
        "page": page, "data": data, "dari": dari, "sampai": sampai, "tab": tab,
        "kategori": kategori,
    })


@login_required
def settlement_pending(request):
    """Settlement Tertunda: antrean baris kredit menunggu uang tiba (H+1),
    dengan umur & sisa jendela. Filter DP/WD, urut tertua dulu."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    rows = pending_settlement_rows(active)
    flow = request.GET.get("flow", "")
    if flow == "depo":
        rows = [r for r in rows if r["jenis"] == "depo"]
    elif flow == "wd":
        rows = [r for r in rows if r["jenis"] == "wd"]
    else:
        flow = ""
    total_nominal = sum((r["nominal"] or 0) for r in rows)
    page = Paginator(rows, 40).get_page(request.GET.get("page"))
    return render(request, "web/settlement.html", {
        "page": page, "flow": flow, "total_count": len(rows),
        "total_nominal": total_nominal,
    })


@login_required
def rekening_breakdown(request):
    """Rincian Rekening: breakdown sisi uang (bank/gateway) per rekening operator
    pada satu tanggal — Deposit/Withdraw/Admin/Net/Saldo + Selisih Kontrol."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    latest = Transaction.objects.filter(
        toko=active, source_type__key__in=("bank", "gateway")
    ).aggregate(m=Max("occurred_at__date"))["m"]
    tanggal = _parse_date(request.GET.get("date", "")) or latest or date_cls.today()
    data = hitung_rekening_breakdown(active, tanggal)
    return render(request, "web/rekening.html", {
        "data": data, "tanggal": tanggal, "latest": latest,
        "prev_date": tanggal - timedelta(days=1),
        "next_date": tanggal + timedelta(days=1),
    })


@login_required
def monthly_overview(request):
    """Ringkasan Bulanan: rekap batch harian toko aktif dalam satu bulan
    (Panel/Uang DP & WD, selisih, bucket) langsung dari ReconBatch.summary."""
    active = _active_toko(request)
    if active is None:
        return render(request, "web/no_toko.html")
    batch_dates = ReconBatch.objects.filter(
        toko=active, recon_date__isnull=False
    ).aggregate(m=Max("recon_date"), n=Min("recon_date"))
    latest = batch_dates["m"]
    sel = request.GET.get("month", "")
    year = month = None
    if len(sel) == 7 and sel[4] == "-":
        try:
            year, month = int(sel[:4]), int(sel[5:7])
        except ValueError:
            year = month = None
    if year is None:
        ref = latest or date_cls.today()
        year, month = ref.year, ref.month
    data = monthly_summary(active, year, month)
    # daftar bulan yang punya batch (untuk dropdown)
    months = sorted(
        {(d.year, d.month) for d in ReconBatch.objects.filter(
            toko=active, recon_date__isnull=False
        ).values_list("recon_date", flat=True) if d},
        reverse=True,
    )
    return render(request, "web/monthly.html", {
        "data": data,
        "bulan": date_cls(year, month, 1),
        "months": [date_cls(y, m, 1) for y, m in months],
        "sel_month": f"{year:04d}-{month:02d}",
    })


@login_required
def toko_overview(request):
    """Ringkasan lintas toko (scoped RBAC): status rekon terakhir, selisih,
    antrean tinjau, menunggu settlement, uang periksa D.
    Filter tanggal (?from&to): selisih/tinjau/uang-D diagregasi atas batch
    DALAM rentang; 'rekon terakhir' = batch terakhir dalam rentang.
    Semua angka diambil lewat query AGREGAT lintas toko (bukan per toko) —
    versi per-toko membuat klik menu Toko ~1,5 dtk di prod (24 toko × 3 query)."""
    from reconciliation.engine import pending_settlement_counts

    date_from = _parse_date(request.GET.get("from", ""))
    date_to = _parse_date(request.GET.get("to", ""))
    filtered = bool(date_from or date_to)
    tokos = list(tokos_for(request.user))

    batch_qs = ReconBatch.objects.filter(toko__in=tokos, recon_date__isnull=False)
    tinjau_qs = MatchResult.objects.filter(
        run__batch__toko__in=tokos, bucket=MatchResult.Bucket.TINJAU
    )
    if date_from:
        batch_qs = batch_qs.filter(recon_date__gte=date_from)
        tinjau_qs = tinjau_qs.filter(run__batch__recon_date__gte=date_from)
    if date_to:
        batch_qs = batch_qs.filter(recon_date__lte=date_to)
        tinjau_qs = tinjau_qs.filter(run__batch__recon_date__lte=date_to)
    batches_by_toko = {}
    for b in batch_qs.order_by("recon_date"):
        batches_by_toko.setdefault(b.toko_id, []).append(b)
    tinjau_by_toko = dict(
        tinjau_qs.values_list("run__batch__toko").annotate(n=Count("id"))
    )
    pending_by_toko = pending_settlement_counts(tokos)

    rows = []
    for t in tokos:
        batches = batches_by_toko.get(t.id, [])
        last = batches[-1] if batches else None
        # tanpa filter: perilaku lama (angka batch TERAKHIR saja);
        # dengan filter: agregat seluruh batch dalam rentang.
        scope = batches if filtered else batches[-1:]
        total = uang_d = 0
        for b in scope:
            s = b.summary or {}
            total += abs((s.get("dp") or {}).get("selisih") or 0)
            total += abs((s.get("wd") or {}).get("selisih") or 0)
            uang_d += ((s.get("unmatched_money") or {}).get("d") or {}).get("n") or 0
        st = "" if last is None else ("ok" if total == 0 else ("warn" if total < 10_000_000 else "bad"))
        rows.append({
            "toko": t, "last": last, "selisih": total, "status": st,
            "tinjau": tinjau_by_toko.get(t.id, 0),
            "pending": pending_by_toko.get(t.id, 0),
            "uang_d": uang_d,
            "has_batch": last is not None,
        })
    # selisih terbesar dulu; toko tanpa batch di bawah
    rows.sort(key=lambda r: (r["has_batch"], r["selisih"]), reverse=True)
    return render(request, "web/toko_overview.html", {
        "rows": rows,
        "date_from": request.GET.get("from", "") if date_from else "",
        "date_to": request.GET.get("to", "") if date_to else "",
    })


def _baca_alasan(request):
    """Kode alasan OPSIONAL aksi review — daftar sah REUSE FRKoreksi.ALASAN_KOREKSI
    (satu sumber kebenaran; tanpa duplikasi list). Return (alasan, valid)."""
    alasan = request.POST.get("alasan") or ""
    if alasan and alasan not in dict(FRKoreksi.ALASAN_KOREKSI):
        return "", False
    return alasan, True


def _annot_alasan_review(qs):
    """Tempelkan alasan+catatan ReviewAction terakhir per hasil (untuk chip
    di baris manual_override) — satu subquery, tanpa N+1."""
    terakhir = ReviewAction.objects.filter(result=OuterRef("pk")).order_by("-id")
    return qs.annotate(
        alasan_manual=Subquery(terakhir.values("alasan")[:1]),
        catatan_manual=Subquery(terakhir.values("reason")[:1]),
    )


@login_required
@require_POST
def bulk_review(request, pk):
    """Setujui / tandai-tinjau banyak hasil sekaligus (per halaman terfilter).
    Setiap baris tetap tercatat ReviewAction-nya sendiri — jejak audit utuh."""
    run = get_object_or_404(MatchRun, pk=pk, batch__toko__in=tokos_for(request.user))
    action = request.POST.get("action", "")
    buckets = {"mark_matched": MatchResult.Bucket.COCOK,
               "mark_review": MatchResult.Bucket.TINJAU}
    if action not in buckets:
        return HttpResponseBadRequest("Aksi tidak dikenal.")
    alasan, sah = _baca_alasan(request)
    if not sah:
        return HttpResponseBadRequest("alasan tidak dikenal")
    catatan = (request.POST.get("catatan") or "").strip()
    ids = [i for i in request.POST.getlist("result_ids") if i.isdigit()]
    rows = list(MatchResult.objects.filter(run=run, id__in=ids))
    for r in rows:
        r.bucket = buckets[action]
        r.reason_code = "manual_override"
        r.save(update_fields=["bucket", "reason_code"])
        ReviewAction.objects.create(
            result=r, action=action, reason=catatan or "bulk",
            alasan=alasan, reviewer=request.user
        )
    if rows:
        catat(request.user, "review_massal", f"{len(rows)} hasil",
              toko=run.batch.toko if run.batch else None,
              run_pk=run.pk, n=len(rows), action=action, alasan=alasan)
        if run.batch:  # kartu Cocok/Tinjau run & batch jangan basi terhadap chip live
            refresh_batch_summary(run.batch)
    messages.success(request, f"{len(rows)} hasil diperbarui.")
    nxt = request.POST.get("next") or reverse("run_detail", args=[run.pk])
    if not url_has_allowed_host_and_scheme(
        nxt, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        nxt = reverse("run_detail", args=[run.pk])
    return redirect(nxt)


@login_required
@require_POST
def bulk_review_queue(request):
    """Setujui / tandai-tinjau massal LINTAS-run dari Area Pengecekan.
    Mutasi per baris identik bulk_review; ringkasan tiap batch tersentuh
    disegarkan sekali."""
    action = request.POST.get("action", "")
    buckets = {"mark_matched": MatchResult.Bucket.COCOK,
               "mark_review": MatchResult.Bucket.TINJAU}
    if action not in buckets:
        return HttpResponseBadRequest("Aksi tidak dikenal.")
    alasan, sah = _baca_alasan(request)
    if not sah:
        return HttpResponseBadRequest("alasan tidak dikenal")
    catatan = (request.POST.get("catatan") or "").strip()
    ids = [i for i in request.POST.getlist("result_ids") if i.isdigit()]
    rows = list(MatchResult.objects.filter(
        id__in=ids, run__batch__toko__in=tokos_for(request.user)
    ).select_related("run__batch"))
    batches = {}
    for r in rows:
        r.bucket = buckets[action]
        r.reason_code = "manual_override"
        r.save(update_fields=["bucket", "reason_code"])
        ReviewAction.objects.create(
            result=r, action=action, reason=catatan or "bulk",
            alasan=alasan, reviewer=request.user)
        if r.run.batch_id:
            batches[r.run.batch_id] = r.run.batch
    if rows:
        catat(request.user, "review_massal", f"{len(rows)} hasil (Area Pengecekan)",
              toko=rows[0].run.batch.toko if rows[0].run.batch else None,
              n=len(rows), action=action, alasan=alasan)
        for b in batches.values():  # kartu run & batch jangan basi
            refresh_batch_summary(b)
    messages.success(request, f"{len(rows)} hasil diperbarui.")
    nxt = request.POST.get("next") or reverse("review_queue")
    if not url_has_allowed_host_and_scheme(
        nxt, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        nxt = reverse("review_queue")
    return redirect(nxt)


@login_required
def review_alasan_form(request, pk):
    """Popup kecil pilih alasan sebelum Setujui/Tinjau satu hasil (GET, HTMX) —
    cetakan sama dgn modal Koreksi FR."""
    r = get_object_or_404(MatchResult, pk=pk, run__batch__toko__in=tokos_for(request.user))
    action = request.GET.get("action", "")
    judul = {"mark_matched": "Setujui hasil", "mark_review": "Tandai perlu ditinjau"}
    if action not in judul:
        return HttpResponseBadRequest("Aksi tidak dikenal.")
    return render(request, "web/_review_alasan_form.html", {
        "r": r, "action": action, "judul": judul[action],
        "show_run_col": request.GET.get("show_run_col") == "1",
        "hide_left": request.GET.get("hide_left") == "1",
        "pilihan_alasan": FRKoreksi.ALASAN_KOREKSI,
    })


@login_required
@require_POST
def review(request, pk):
    r = get_object_or_404(MatchResult, pk=pk, run__batch__toko__in=tokos_for(request.user))
    action = request.POST.get("action", "")
    buckets = {
        "mark_matched": MatchResult.Bucket.COCOK,
        "mark_review": MatchResult.Bucket.TINJAU,
        "mark_unmatched": MatchResult.Bucket.TIDAK,
    }
    if action not in buckets:
        return HttpResponseBadRequest("Aksi tidak dikenal.")
    alasan, sah = _baca_alasan(request)
    if not sah:
        return HttpResponseBadRequest("alasan tidak dikenal")
    # catatan bebas: param baru `catatan` (modal alasan) ATAU param lama `reason`
    reason = (request.POST.get("catatan") or "").strip() or request.POST.get("reason", "")
    # Catatan: override pada hasil no_money yang barisnya masih AKTIF (menunggu
    # settlement) mengeluarkannya dari carry-over — baris itu akan diperlakukan
    # sebagai baris baru di run berikutnya. Follow-up kecil bila jadi masalah:
    # konsumsi baris ke batch asalnya saat di-override.
    r.bucket = buckets[action]
    r.reason_code = "manual_override"
    r.save(update_fields=["bucket", "reason_code"])
    ReviewAction.objects.create(
        result=r, action=action, reason=reason, alasan=alasan, reviewer=request.user
    )
    catat(request.user, "review", f"Result #{r.pk}",
          toko=r.run.batch.toko if r.run.batch else None, result_pk=r.pk, action=action,
          alasan=alasan)
    if r.run.batch:  # kartu Cocok/Tinjau run & batch jangan basi terhadap chip live
        refresh_batch_summary(r.run.batch)
    show_run_col = request.POST.get("show_run_col") == "1"
    if show_run_col and r.run.batch:
        r.home_no = ReconBatch.objects.filter(
            toko=r.run.batch.toko, id__lte=r.run.batch_id
        ).count()
    # chip alasan harus langsung tampil di baris hasil swap (tanpa annotate queryset)
    r.alasan_manual = alasan
    r.catatan_manual = reason
    html = render_to_string("web/_result_row.html", {
        "r": r, "bucket_meta": BUCKET_META, "show_run_col": show_run_col,
        "hide_left": request.POST.get("hide_left") == "1",
    }, request=request)
    # tutup modal alasan (bila aksi datang dari popup) — div OOB di belakang fragmen
    # <tr>. BUTUH htmx useTemplateFragments=true (meta htmx-config di app_base):
    # tanpa itu parser legacy membungkus respons <tr> dgn <table><tbody> dan div
    # non-tabel ini di-foster-parent keluar -> fragmen kosong (baris lenyap, modal
    # tak tertutup). Beda dgn fr_koreksi_simpan yang fragmen utamanya <div>.
    html += '<div id="reviewPop" hx-swap-oob="innerHTML"></div>'
    return HttpResponse(html)


@login_required
def export_run(request, pk):
    import io

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font

    run = get_object_or_404(MatchRun, pk=pk, batch__toko__in=tokos_for(request.user))
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    s = run.summary or {}
    for label, val in [
        ("Rekonsiliasi", run.get_relation_display()),
        ("Run", f"#{run.pk}"),
        ("Toleransi", f"{run.tolerance.name} (±{run.tolerance.date_window_days} hari)"),
        ("Tanggal", run.created_at.strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Cocok", s.get("cocok", 0)),
        ("Perlu Ditinjau", s.get("perlu_tinjau", 0)),
        ("Tidak Cocok", s.get("tidak_cocok", 0)),
    ]:
        ws.append([label, val])
    for row in ws["A"]:
        row.font = Font(bold=True)

    from web.exports import results_sheet

    results_sheet(wb, run, "Hasil", REL_LABELS)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="rekonsiliasi_run{run.pk}.xlsx"'
    return resp


# Batas jumlah batch satu kali export bulk — jaga memori & waktu respons.
EXPORT_BATCH_CAP = 200


@login_required
def export_center(request):
    """Menu Export: per-tanggal / per-toko / semua toko (admin) / kombinasi.

    1 batch -> xlsx langsung; >1 -> ZIP berisi xlsx per-(toko,tanggal) —
    output tetap per-tanggal walaupun export dilakukan sekaligus (bulk).
    """
    import io

    from web.exports import XLSX_CT, batch_filename, build_batch_workbook, safe_name

    allowed = tokos_for(request.user)
    toko_param = request.GET.get("toko", "")
    date_from = _parse_date(request.GET.get("from", ""))
    date_to = _parse_date(request.GET.get("to", ""))
    # UX rentang: isi "Dari" saja = tarik satu tanggal (Sampai ikut); isi "Sampai"
    # saja = anggap tanggal itu juga. Isi keduanya = rentang.
    if date_from and not date_to:
        date_to = date_from
    elif date_to and not date_from:
        date_from = date_to

    if not toko_param:  # form
        return render(request, "web/export_center.html", {"tokos": allowed})

    # --- resolusi scope toko (RBAC: tokos_for = satu-satunya sumber kebenaran) ---
    if toko_param == "all":
        if not is_admin(request.user):
            messages.error(request, "Export semua toko khusus admin.")
            return redirect("export_center")
        scope = allowed
        scope_label = "semua-toko"
    else:
        toko = allowed.filter(id=toko_param).first() if toko_param.isdigit() else None
        if toko is None:
            messages.error(request, "Toko tidak dikenal atau bukan wewenangmu.")
            return redirect("export_center")
        scope = [toko]
        scope_label = safe_name(toko.name)

    batches = (
        ReconBatch.objects.filter(toko__in=scope, recon_date__isnull=False)
        .select_related("toko", "tolerance")
        .order_by("toko__name", "recon_date")
    )
    if date_from:
        batches = batches.filter(recon_date__gte=date_from)
    if date_to:
        batches = batches.filter(recon_date__lte=date_to)

    n = batches.count()
    if n == 0:
        messages.error(request, "Tidak ada batch pada pilihan itu — cek toko/tanggalnya.")
        return redirect("export_center")
    if n > EXPORT_BATCH_CAP:
        messages.error(
            request, f"{n} batch terlalu banyak untuk satu export — persempit rentang tanggal (maks {EXPORT_BATCH_CAP})."
        )
        return redirect("export_center")

    # nomor batch per-toko (konsisten dgn batch_detail) tanpa N query per batch
    def batch_no(b):
        return ReconBatch.objects.filter(toko=b.toko, id__lte=b.id).count()

    catat(request.user, "export_batch", f"{n} batch ({scope_label})")

    if n == 1:
        b = batches[0]
        wb = build_batch_workbook(b, batch_no(b), REL_LABELS)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type=XLSX_CT)
        resp["Content-Disposition"] = f'attachment; filename="{batch_filename(b)}"'
        return resp

    # Bulk: ZIP — satu xlsx per (toko, tanggal); workbook dibangun berurutan (hemat memori).
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        for b in batches.iterator():
            wb = build_batch_workbook(b, batch_no(b), REL_LABELS)
            inner = io.BytesIO()
            wb.save(inner)
            zf.writestr(batch_filename(b), inner.getvalue())
    zbuf.seek(0)
    tag_from = date_from.isoformat() if date_from else "awal"
    tag_to = date_to.isoformat() if date_to else "akhir"
    resp = HttpResponse(zbuf.read(), content_type="application/zip")
    resp["Content-Disposition"] = (
        f'attachment; filename="rekonsiliasi_{scope_label}_{tag_from}_{tag_to}.zip"'
    )
    return resp

"""Penyesuaian Export (I3).

(A) Workbook batch diperkaya sheet "Breakdown Bracket" + "Rincian Rekening" bila
    tanggal batch punya data FR/rekening — sheet lama (Ringkasan/Hasil) tak regresi.
(B) Halaman /bracket/export/ mengekspor tabel breakdown SESUAI filter aktif
    (rentang Dari–Sampai + carry-forward saldo dari I1), di-scope toko aktif.
"""
import io
from datetime import date, datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.models import AuditLog
from reconciliation.models import MatchResult, MatchRun, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction
from web.breakdown import bracket_breakdown
from web.exports import (
    XLSX_CT,
    _num,
    breakdown_sheet,
    build_batch_workbook,
    rekening_sheet,
)
from web.rekening import rekening_breakdown
from web.views import REL_LABELS

User = get_user_model()
_seq = iter(range(1, 1_000_000))

D = date(2026, 7, 1)


def _header_index(ws):
    """Indeks (1-based) baris header — baris yang sel pertamanya 'No'."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "No":
            return i
    raise AssertionError("baris header 'No' tidak ditemukan")


def _row_by_second_col(ws, needle):
    """Baris (tuple nilai) yang sel ke-2 (kolom nama) memuat `needle`."""
    for row in ws.iter_rows(values_only=True):
        if len(row) > 1 and row[1] and needle in str(row[1]):
            return row
    return None


def _val(ws, row, colname):
    """Nilai sel `row` pada kolom berjudul `colname` (kebal jumlah kolom kategori)."""
    header = [c.value for c in ws[_header_index(ws)]]
    return row[header.index(colname)]


class _Base(TestCase):
    def setUp(self):
        self.toko = Toko.objects.get(key="lbs")
        self.slo = Toko.objects.get(key="slo")
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.bracket = SourceType.objects.get_or_create(
            key="bracket", defaults={"name": "Bracket"}
        )[0]
        self.bank = SourceType.objects.get_or_create(key="bank", defaults={"name": "Bank"})[0]
        self.gateway = SourceType.objects.get_or_create(
            key="gateway", defaults={"name": "Gateway"}
        )[0]
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.up_br = Upload.objects.create(source_type=self.bracket, toko=self.toko)

    def fr(self, bank, kategori, total, saldo, jam="10:00", tanggal=D, toko=None, up=None):
        toko = toko or self.toko
        up = up or self.up_br
        n = next(_seq)
        return Transaction.objects.create(
            upload=up, source_type=self.bracket, toko=toko,
            jenis="lainnya", amount=abs(Decimal(total)), money_delta=Decimal(total),
            balance_after=None if saldo is None else Decimal(saldo),
            posted_date=tanggal, occurred_at=datetime(tanggal.year, tanggal.month, tanggal.day, 10, 0),
            row_hash=f"br{n}", raw={"Bank": bank, "Kategori": kategori, "Jam": jam},
        )

    def mutasi(self, provider, owner, delta, saldo, tanggal=D, jenis="lainnya", gateway=False):
        """Satu baris mutasi bank/gateway → satu rekening `PROVIDER a/n OWNER`."""
        st = self.gateway if gateway else self.bank
        up = Upload.objects.create(
            source_type=st, toko=self.toko, provider=provider, owner_name=owner
        )
        n = next(_seq)
        return Transaction.objects.create(
            upload=up, source_type=st, toko=self.toko, jenis=jenis,
            amount=abs(Decimal(delta)), money_delta=Decimal(delta),
            balance_after=None if saldo is None else Decimal(saldo),
            occurred_at=datetime(tanggal.year, tanggal.month, tanggal.day, 11, 0),
            row_hash=f"bk{n}", raw={},
        )

    def _batch(self, toko=None, recon_date=D):
        toko = toko or self.toko
        batch = ReconBatch.objects.create(
            toko=toko, tolerance=self.tol, recon_date=recon_date,
            summary={"dp": {"panel": 100, "money_matched": 100, "selisih": 0},
                     "wd": {"panel": 0, "money_matched": 0, "selisih": 0},
                     "buckets": {"cocok": 1, "perlu_tinjau": 0, "tidak_cocok": 0}},
        )
        run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=batch,
            summary={"left": 1, "cocok": 1, "perlu_tinjau": 0, "tidak_cocok": 0},
        )
        up = Upload.objects.create(source_type=self.panel, toko=toko)
        left = Transaction.objects.create(
            upload=up, source_type=self.panel, toko=toko, jenis="depo",
            amount=Decimal("50000"), ticket_no=f"D-{next(_seq)}", row_hash=f"pn{next(_seq)}", raw={},
        )
        MatchResult.objects.create(
            run=run, bucket=MatchResult.Bucket.COCOK, left=left, reason_code="ticket_exact"
        )
        return batch


# --------------------------------------------------------------------------- #
#  Helper _num + sheet builder (unit — tanpa view)                            #
# --------------------------------------------------------------------------- #
class NumHelperTests(TestCase):
    def test_none_jadi_string_kosong(self):
        self.assertEqual(_num(None), "")

    def test_decimal_jadi_float(self):
        self.assertEqual(_num(Decimal("500000")), 500000.0)
        self.assertIsInstance(_num(Decimal("500000")), float)


class BreakdownSheetTests(_Base):
    def test_header_kolom_dinamis_dan_caption(self):
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000", jam="09:00")
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "BEBAN ADMIN QRIS", "-4972", "1495028", jam="10:30")
        data = bracket_breakdown(self.toko, D)
        wb = Workbook()
        ws = breakdown_sheet(wb, data, "Breakdown Bracket", "01/07/2026")
        # caption berisi label tanggal, header di baris berikutnya
        self.assertEqual(ws["A1"].value, "01/07/2026")
        hidx = _header_index(ws)
        header = [c.value for c in ws[hidx]]
        self.assertEqual(header, [
            "No", "FR Account", "Saldo Awal", "Deposit", "Beban Admin QRIS",
            "Total Mutasi", "Saldo Akhir", "Selisih Kontrol",
        ])

    def test_nilai_saldo_dan_selisih_benar(self):
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000", jam="09:00")
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "BEBAN ADMIN QRIS", "-4972", "1495028", jam="10:30")
        data = bracket_breakdown(self.toko, D)
        wb = Workbook()
        ws = breakdown_sheet(wb, data, "BD", "01/07/2026")
        row = _row_by_second_col(ws, "QRIS HOKI")
        self.assertIsNotNone(row)
        # No, FR Account, Saldo Awal, Deposit, Beban Admin QRIS, Total Mutasi, Saldo Akhir, Selisih
        self.assertEqual(row[2], 1000000.0)          # saldo awal
        self.assertEqual(row[3], 500000.0)           # deposit
        self.assertEqual(row[4], -4972.0)            # beban admin qris
        self.assertEqual(row[5], 495028.0)           # total mutasi
        self.assertEqual(row[6], 1495028.0)          # saldo akhir
        self.assertEqual(row[7], 0.0)                # selisih kontrol

    def test_total_row_ada(self):
        self.fr("BANK BCA | IRFAN | DEPOSIT", "Deposit", "100000", "150000", jam="09:00")
        data = bracket_breakdown(self.toko, D)
        wb = Workbook()
        ws = breakdown_sheet(wb, data, "BD", "01/07/2026")
        row = _row_by_second_col(ws, "TOTAL")
        self.assertIsNotNone(row)

    def test_saldo_none_jadi_sel_kosong(self):
        # akun tanpa balance → saldo awal/akhir/selisih None → sel kosong (bukan 0)
        self.fr("", "Adjustment", "1000", None)
        data = bracket_breakdown(self.toko, D)
        wb = Workbook()
        ws = breakdown_sheet(wb, data, "BD", "01/07/2026")
        row = _row_by_second_col(ws, "Tanpa Akun")
        self.assertIsNotNone(row)
        self.assertEqual(row[2], "")   # saldo awal kosong
        self.assertEqual(row[-1], "")  # selisih kontrol kosong


class RekeningSheetTests(_Base):
    def test_header_memuat_biaya_admin(self):
        self.mutasi("BCA", "HENDI", "100000", "600000")
        data = rekening_breakdown(self.toko, D)
        wb = Workbook()
        ws = rekening_sheet(wb, data, "Rincian Rekening")
        header = [c.value for c in ws[1]]
        self.assertEqual(header, [
            "No", "Rekening", "Deposit", "Withdraw", "Biaya Admin",
            "Net", "Trx", "Saldo Awal", "Saldo Akhir", "Selisih Kontrol",
        ])

    def test_nilai_deposit_dan_admin(self):
        self.mutasi("BCA", "HENDI", "100000", "600000")
        self.mutasi("BCA", "HENDI", "-1000", "599000", jenis="admin")
        data = rekening_breakdown(self.toko, D)
        wb = Workbook()
        ws = rekening_sheet(wb, data, "R")
        row = _row_by_second_col(ws, "BCA a/n HENDI")
        self.assertIsNotNone(row)
        self.assertEqual(row[2], 100000.0)   # deposit
        self.assertEqual(row[4], -1000.0)    # biaya admin: delta bertanda (debit −1000) apa adanya

    def test_gateway_tanpa_saldo_sel_kosong(self):
        self.mutasi("QRIS", "HOKI", "50000", None, gateway=True)
        data = rekening_breakdown(self.toko, D)
        wb = Workbook()
        ws = rekening_sheet(wb, data, "R")
        row = _row_by_second_col(ws, "QRIS")
        self.assertIsNotNone(row)
        self.assertEqual(row[7], "")   # saldo awal kosong
        self.assertEqual(row[8], "")   # saldo akhir kosong
        self.assertEqual(row[9], "")   # selisih kontrol kosong


# --------------------------------------------------------------------------- #
#  (A) build_batch_workbook diperkaya                                         #
# --------------------------------------------------------------------------- #
class BatchWorkbookTests(_Base):
    def test_tambah_sheet_breakdown_dan_rekening(self):
        batch = self._batch()
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000")
        self.mutasi("BCA", "HENDI", "100000", "600000")
        wb = build_batch_workbook(batch, 1, REL_LABELS)
        names = wb.sheetnames
        self.assertIn("Ringkasan", names)
        self.assertTrue(any(s.startswith("Hasil") for s in names))
        self.assertIn("Breakdown Bracket", names)
        self.assertIn("Rincian Rekening", names)

    def test_tanpa_data_fr_tetap_kompat(self):
        # batch panel-only → tak ada baris FR/rekening pada tanggalnya → sheet
        # breakdown/rekening TIDAK ditambahkan (jaga kompat export_center lama).
        batch = self._batch()
        wb = build_batch_workbook(batch, 1, REL_LABELS)
        names = set(wb.sheetnames)
        self.assertIn("Ringkasan", names)
        self.assertTrue(any(s.startswith("Hasil") for s in names))
        self.assertNotIn("Breakdown Bracket", names)
        self.assertNotIn("Rincian Rekening", names)

    def test_data_tanggal_lain_tak_bikin_sheet(self):
        # data FR ada tapi di tanggal ≠ recon_date → tak menambah sheet breakdown.
        batch = self._batch(recon_date=D)
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000",
                tanggal=date(2026, 7, 5))
        wb = build_batch_workbook(batch, 1, REL_LABELS)
        self.assertNotIn("Breakdown Bracket", wb.sheetnames)


# --------------------------------------------------------------------------- #
#  (B) view export_breakdown                                                  #
# --------------------------------------------------------------------------- #
class ExportBreakdownViewTests(_Base):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user("sup", "s@a.co", "pw12345", role="supervisor")
        self.client.login(username="sup", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})

    def test_butuh_login(self):
        self.client.logout()
        r = self.client.get(reverse("export_breakdown"))
        self.assertEqual(r.status_code, 302)
        self.assertIn("login", r["Location"])

    def test_xlsx_dan_nama_file(self):
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000")
        r = self.client.get(reverse("export_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-01"})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], XLSX_CT)
        cd = r["Content-Disposition"]
        self.assertIn("breakdown_", cd)
        self.assertIn(self.toko.name.replace(" ", "_"), cd)
        self.assertIn("2026-07-01", cd)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertIn("Breakdown Bracket", wb.sheetnames)

    def test_hormati_rentang(self):
        self.fr("BANK BCA | IRFAN | DEPOSIT", "Deposit", "100000", "100000", tanggal=date(2026, 7, 1))
        self.fr("QRIS FLYER | DEPOSIT / WITHDRAW", "Deposit", "90000", "90000", tanggal=date(2026, 7, 2))
        # rentang penuh → kedua akun ikut
        r = self.client.get(reverse("export_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-02"})
        ws = load_workbook(io.BytesIO(r.content)).active
        self.assertIsNotNone(_row_by_second_col(ws, "BANK BCA"))
        self.assertIsNotNone(_row_by_second_col(ws, "QRIS FLYER"))
        # hanya 01 → akun 02 TIDAK ikut
        r1 = self.client.get(reverse("export_breakdown"),
                             {"dari": "2026-07-01", "sampai": "2026-07-01"})
        ws1 = load_workbook(io.BytesIO(r1.content)).active
        self.assertIsNotNone(_row_by_second_col(ws1, "BANK BCA"))
        self.assertIsNone(_row_by_second_col(ws1, "QRIS FLYER"))

    def test_carry_forward_saldo_awal(self):
        # akun dorman ber-saldo di H-1 → tampil di export H dengan saldo_awal = penutup
        self.fr("BANK BRI | YOGA | WITHDRAW", "Deposit", "200000", "500000",
                tanggal=date(2026, 6, 30))
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "50000", "50000",
                tanggal=date(2026, 7, 1))
        r = self.client.get(reverse("export_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-01"})
        ws = load_workbook(io.BytesIO(r.content)).active
        row = _row_by_second_col(ws, "BANK BRI")
        self.assertIsNotNone(row)          # akun dorman ikut (carry-forward)
        self.assertEqual(_val(ws, row, "Saldo Awal"), 500000.0)   # penutup H-1
        self.assertEqual(_val(ws, row, "Total Mutasi"), 0.0)      # tak ada gerak di H

    def test_koreksi_terpakai_pada_satu_hari(self):
        from web.models import FRKoreksi
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000")
        FRKoreksi.objects.create(
            toko=self.toko, tanggal=D, account="QRIS HOKI | DEPOSIT / WITHDRAW",
            kolom="deposit", nilai=Decimal("777000"), dibuat_oleh=self.user,
        )
        r = self.client.get(reverse("export_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-01"})
        ws = load_workbook(io.BytesIO(r.content)).active
        row = _row_by_second_col(ws, "QRIS HOKI")
        # kolom Deposit (indeks 3) memakai nilai terkoreksi
        self.assertEqual(row[3], 777000.0)

    def test_scope_rbac_toko_lain_tak_bocor(self):
        # data FR untuk SLO; auditor hanya berwenang LBS → export tak memuat SLO
        up_slo = Upload.objects.create(source_type=self.bracket, toko=self.slo)
        self.fr("BANK SLO | X | DEPOSIT", "Deposit", "123456", "123456",
                toko=self.slo, up=up_slo)
        self.client.logout()
        aud = User.objects.create_user("aud", "a@a.co", "pw12345", role="auditor")
        aud.allowed_tokos.set([self.toko])
        self.client.login(username="aud", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})
        r = self.client.get(reverse("export_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-01"})
        self.assertEqual(r["Content-Type"], XLSX_CT)
        self.assertNotIn(b"BANK SLO", r.content)

    def test_audit_tercatat(self):
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000")
        self.client.get(reverse("export_breakdown"),
                        {"dari": "2026-07-01", "sampai": "2026-07-01"})
        self.assertTrue(AuditLog.objects.filter(aksi="export_breakdown").exists())


class ExportBreakdownButtonTests(_Base):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user("sup2", "s2@a.co", "pw12345", role="supervisor")
        self.client.login(username="sup2", password="pw12345")
        self.client.post(reverse("set_toko"), {"toko_id": self.toko.id})

    def test_tombol_export_di_halaman_breakdown(self):
        self.fr("QRIS HOKI | DEPOSIT / WITHDRAW", "Deposit", "500000", "1500000")
        r = self.client.get(reverse("bracket_breakdown"),
                            {"dari": "2026-07-01", "sampai": "2026-07-01"})
        html = r.content.decode()
        self.assertIn(reverse("export_breakdown"), html)
        # tombol membawa rentang aktif
        self.assertIn("dari=2026-07-01", html)
        self.assertIn("sampai=2026-07-01", html)

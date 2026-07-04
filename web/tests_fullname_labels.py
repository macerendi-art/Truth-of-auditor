"""Test Task 3: boost nama lengkap di engine + label sumber asli (bukan Kiri/Kanan)."""
import io
from datetime import datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from reconciliation.engine import run_match
from reconciliation.models import MatchResult, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction


class MoneyMatcherFullNameBoostTests(TestCase):
    """Cabang username&username: boost via counterparty (Full Name) bila username beda."""

    def setUp(self):
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.lbs = Toko.objects.get(key="lbs")
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.gateway = SourceType.objects.get_or_create(key="gateway", defaults={"name": "Gateway"})[0]
        self.bank = SourceType.objects.get_or_create(key="bank", defaults={"name": "Bank"})[0]
        self.up = Upload.objects.create(source_type=self.panel, toko=self.lbs)
        self.batch = ReconBatch.objects.create(toko=self.lbs, tolerance=self.tol)

    def test_username_beda_tapi_counterparty_identik_tetap_cocok(self):
        """Boost Full Name di jalur BANK fuzzy: username beda -> skor 40 -> tapi Full Name
        identik naikkan skor -> cocok. (Gateway kini via TXN ID eksak, jadi boost fuzzy
        ini berlaku untuk BANK yang tak punya kunci.)"""
        Transaction.objects.create(
            upload=self.up, source_type=self.panel, toko=self.lbs, jenis="depo",
            amount=Decimal("50000"), money_delta=Decimal("50000"),
            occurred_at=datetime(2026, 6, 27, 10, 0),
            username="budi123", counterparty="BUDI SANTOSO",
            row_hash="fn1",
        )
        Transaction.objects.create(
            upload=self.up, source_type=self.bank, toko=self.lbs, jenis="lainnya",
            amount=Decimal("50000"), money_delta=Decimal("50000"),
            occurred_at=datetime(2026, 6, 27, 10, 30),
            username="nx_9981", counterparty="BUDI SANTOSO",
            row_hash="fn2",
        )
        run = run_match("panel_bank", self.tol, user=None, toko=self.lbs, batch=self.batch)
        result = MatchResult.objects.get(run=run)
        self.assertEqual(result.bucket, MatchResult.Bucket.COCOK, f"score={result.score}")
        self.assertGreaterEqual(result.score, 85)

    def test_regresi_bank_tanpa_username_counterparty_mirip_tetap_cocok(self):
        """Cabang else (fuzzy lama) tidak boleh berubah perilakunya."""
        Transaction.objects.create(
            upload=self.up, source_type=self.panel, toko=self.lbs, jenis="depo",
            amount=Decimal("75000"), money_delta=Decimal("75000"),
            occurred_at=datetime(2026, 6, 27, 9, 0),
            username="siti456", counterparty="SITI AMINAH",
            row_hash="fn3",
        )
        Transaction.objects.create(
            upload=self.up, source_type=self.bank, toko=self.lbs, jenis="lainnya",
            amount=Decimal("75000"), money_delta=Decimal("75000"),
            occurred_at=datetime(2026, 6, 27, 9, 15),
            username="", counterparty="SITI AMINAH",
            row_hash="fn4",
        )
        run = run_match("panel_bank", self.tol, user=None, toko=self.lbs, batch=self.batch)
        result = MatchResult.objects.get(run=run)
        self.assertEqual(result.bucket, MatchResult.Bucket.COCOK, f"score={result.score}")


class RunDetailLabelTests(TestCase):
    """Label sumber asli per relasi (bukan hardcode Kiri/Kanan)."""

    def setUp(self):
        User = get_user_model()
        User.objects.create_user("aud", "a@a.co", "pw12345", role="supervisor")
        self.client.login(username="aud", password="pw12345")
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.lbs = Toko.objects.get(key="lbs")
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.gateway = SourceType.objects.get_or_create(key="gateway", defaults={"name": "Gateway"})[0]
        self.up = Upload.objects.create(source_type=self.panel, toko=self.lbs)
        self.batch = ReconBatch.objects.create(toko=self.lbs, tolerance=self.tol)
        Transaction.objects.create(
            upload=self.up, source_type=self.panel, toko=self.lbs, jenis="depo",
            amount=Decimal("60000"), money_delta=Decimal("60000"),
            occurred_at=datetime(2026, 6, 27, 10, 0),
            ticket_no="D0012345", username="dedi77", counterparty="DEDI KURNIAWAN",
            row_hash="fn5",
        )
        Transaction.objects.create(
            upload=self.up, source_type=self.gateway, toko=self.lbs, jenis="lainnya",
            amount=Decimal("60000"), money_delta=Decimal("60000"),
            occurred_at=datetime(2026, 6, 27, 10, 10),
            ticket_no="NX-772211", username="dedi77", counterparty="D KURNIAWAN",
            row_hash="fn6",
        )
        self.client.post(reverse("set_toko"), {"toko_id": self.lbs.id})
        self.run = run_match("panel_bank", self.tol, user=None, toko=self.lbs, batch=self.batch)

    def test_run_detail_pakai_label_sumber_asli_bukan_kiri_panel(self):
        r = self.client.get(reverse("run_detail", args=[self.run.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Bank/Gateway")
        self.assertEqual(r.context["left_label"], "Panel")
        self.assertNotContains(r, "Kiri (Panel)")

    def test_baris_hasil_tampilkan_counterparty_panel(self):
        r = self.client.get(reverse("run_detail", args=[self.run.pk]))
        self.assertContains(r, "DEDI KURNIAWAN")


class ExportRunFullNameTests(TestCase):
    """Export Excel: header 'Panel Nama Lengkap' + data nama lengkap sisi kiri."""

    def setUp(self):
        User = get_user_model()
        User.objects.create_user("aud", "a@a.co", "pw12345", role="supervisor")
        self.client.login(username="aud", password="pw12345")
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.lbs = Toko.objects.get(key="lbs")
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]
        self.gateway = SourceType.objects.get_or_create(key="gateway", defaults={"name": "Gateway"})[0]
        self.up = Upload.objects.create(source_type=self.panel, toko=self.lbs)
        self.batch = ReconBatch.objects.create(toko=self.lbs, tolerance=self.tol)
        Transaction.objects.create(
            upload=self.up, source_type=self.panel, toko=self.lbs, jenis="depo",
            amount=Decimal("45000"), money_delta=Decimal("45000"),
            occurred_at=datetime(2026, 6, 27, 11, 0),
            username="rina88", counterparty="RINA WATI",
            row_hash="fn7",
        )
        Transaction.objects.create(
            upload=self.up, source_type=self.gateway, toko=self.lbs, jenis="lainnya",
            amount=Decimal("45000"), money_delta=Decimal("45000"),
            occurred_at=datetime(2026, 6, 27, 11, 5),
            username="rina88", counterparty="RINA WATI",
            row_hash="fn8",
        )
        self.client.post(reverse("set_toko"), {"toko_id": self.lbs.id})
        self.run = run_match("panel_bank", self.tol, user=None, toko=self.lbs, batch=self.batch)

    def test_export_header_dan_data_nama_lengkap(self):
        r = self.client.get(reverse("export_run", args=[self.run.pk]))
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Hasil"]
        headers = [c.value for c in ws[1]]
        self.assertIn("Panel Nama Lengkap", headers)
        col = headers.index("Panel Nama Lengkap")
        values = [row[col].value for row in ws.iter_rows(min_row=2)]
        self.assertIn("RINA WATI", values)

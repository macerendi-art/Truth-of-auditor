"""Builder workbook export rekonsiliasi — dipakai export_run & export_center.

Satu jalur kode untuk sheet "Hasil" supaya export per-run dan per-batch
tidak pernah beda format.
"""
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font

from core.version import versi as app_versi
from web.templatetags.web_extras import reason_label

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Label relasi utk header kolom — di-inject dari views (hindari import melingkar).
_REL_LABELS_FALLBACK = ("Kiri", "Kanan")


def safe_name(s):
    """Nama file aman lintas OS: selain [A-Za-z0-9-] jadi '_'."""
    return re.sub(r"[^A-Za-z0-9-]+", "_", str(s or "")).strip("_")


def batch_filename(batch):
    """rekonsiliasi_<toko>_<tanggal>.xlsx — permintaan UAT: tanggal + nama toko."""
    toko = safe_name(batch.toko.name if batch.toko else "toko")
    tgl = batch.recon_date.isoformat() if batch.recon_date else f"batch{batch.pk}"
    return f"rekonsiliasi_{toko}_{tgl}.xlsx"


def monthly_filename(toko, year, month):
    """ringkasan_bulanan_<toko>_YYYY-MM.xlsx — export summary 1 bulan."""
    name = safe_name(toko.name if getattr(toko, "name", None) else toko)
    return f"ringkasan_bulanan_{name}_{year:04d}-{month:02d}.xlsx"


def build_monthly_workbook(toko, year, month, data=None):
    """Workbook ringkasan bulanan: 1 baris per tanggal + TOTAL (sama UI Ringkasan Bulanan).

    `data` opsional = keluaran `web.monthly.monthly_summary` — dihitung di sini
    bila tidak diisi supaya view & tes builder bisa memanggil mandiri.
    """
    if data is None:
        from web.monthly import monthly_summary

        data = monthly_summary(toko, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan Bulanan"

    periode = f"{month:02d}/{year:04d}"
    meta = [
        ("Toko", toko.name if getattr(toko, "name", None) else str(toko)),
        ("Periode", periode),
        ("Jenis", "Ringkasan bulanan (agregat batch harian)"),
        ("Versi aplikasi", f"v{app_versi()}"),
        ("", ""),
    ]
    for label, val in meta:
        ws.append([label, val])
    for cell in ws["A"]:
        cell.font = Font(bold=True)

    headers = [
        "Tanggal",
        "Panel DP", "Uang DP", "Selisih DP",
        "Panel WD", "Uang WD", "Selisih WD",
        "Cocok", "Tinjau", "Tidak",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True)

    for r in data.get("rows") or ():
        tgl = r.get("date")
        ws.append([
            tgl.strftime("%d/%m/%Y") if tgl else "",
            _num(r.get("dp_panel", 0)),
            _num(r.get("dp_uang", 0)),
            _num(r.get("dp_selisih", 0)),
            _num(r.get("wd_panel", 0)),
            _num(r.get("wd_uang", 0)),
            _num(r.get("wd_selisih", 0)),
            r.get("cocok", 0) or 0,
            r.get("tinjau", 0) or 0,
            r.get("tidak", 0) or 0,
        ])

    tot = data.get("total") or {}
    ws.append([
        "TOTAL",
        _num(tot.get("dp_panel", 0)),
        _num(tot.get("dp_uang", 0)),
        _num(tot.get("dp_selisih", 0)),
        _num(tot.get("wd_panel", 0)),
        _num(tot.get("wd_uang", 0)),
        _num(tot.get("wd_selisih", 0)),
        tot.get("cocok", 0) or 0,
        tot.get("tinjau", 0) or 0,
        tot.get("tidak", 0) or 0,
    ])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    return wb


def _num(x):
    """Nilai numerik untuk sel Excel: Decimal→float, None→sel kosong (bukan 0).

    Sumber tanpa saldo (gateway QRIS, PDF) menaruh None di saldo/selisih; sel
    kosong menjaga "—" tampilan tak berubah jadi angka palsu.
    """
    return "" if x is None else float(x)


def _akun_label(a):
    """Label FR Account untuk sel: nama + peran (cerminan badge di tabel web)."""
    return f"{a['name']} ({a['role']})" if a.get("role") else a["name"]


def breakdown_sheet(wb, data, title, tanggal_label=""):
    """Sheet Control Bracket (per FR Account) — kolom kategori dinamis.

    Header Indonesia, angka float (Excel yang memformat ribuan), header tebal,
    Saldo Awal/Akhir + Selisih Kontrol, dan baris TOTAL. `tanggal_label` = baris
    kepala keterangan tanggal/rentang breakdown. `data` = keluaran
    `web.breakdown.bracket_breakdown` (punya kunci "kolom", "accounts", "total").
    """
    d = wb.create_sheet(title)
    kolom = data["kolom"]  # [(slug, label)]
    if tanggal_label:
        d.append([tanggal_label])
        d["A1"].font = Font(bold=True)
    header = (
        ["No", "FR Account", "Saldo Awal"]
        + [lbl for _slug, lbl in kolom]
        + ["Total Mutasi", "Saldo Akhir", "Selisih Kontrol"]
    )
    d.append(header)
    for c in d[d.max_row]:
        c.font = Font(bold=True)
    for i, a in enumerate(data["accounts"], 1):
        d.append(
            [i, _akun_label(a), _num(a["saldo_awal"])]
            + [_num(a["kategori"].get(slug)) for slug, _lbl in kolom]
            + [_num(a["mutasi"]), _num(a["saldo_akhir"]), _num(a["selisih"])]
        )
    total = data["total"]
    d.append(
        ["", "TOTAL", _num(total["saldo_awal"])]
        + [_num(total["kategori"].get(slug)) for slug, _lbl in kolom]
        + [_num(total["mutasi"]), _num(total["saldo_akhir"]), _num(total["selisih"])]
    )
    for c in d[d.max_row]:
        c.font = Font(bold=True)
    return d


def rekening_sheet(wb, data, title):
    """Sheet Rincian Rekening (per rekening bank/gateway) — header Indonesia + TOTAL.

    `data` = keluaran `web.rekening.rekening_breakdown`.
    """
    d = wb.create_sheet(title)
    header = ["No", "Rekening", "Deposit", "Withdraw", "Biaya Admin",
              "Net", "Trx", "Saldo Awal", "Saldo Akhir", "Selisih Kontrol"]
    d.append(header)
    for c in d[1]:
        c.font = Font(bold=True)
    for i, a in enumerate(data["accounts"], 1):
        label = a["label"] + (" (GATEWAY)" if a.get("is_gateway") else "")
        d.append([
            i, label, _num(a["deposit"]), _num(a["withdraw"]), _num(a["admin"]),
            _num(a["net"]), a["trx"], _num(a["saldo_awal"]),
            _num(a["saldo_akhir"]), _num(a["selisih"]),
        ])
    total = data["total"]
    d.append([
        "", "TOTAL", _num(total["deposit"]), _num(total["withdraw"]), _num(total["admin"]),
        _num(total["net"]), total["trx"], _num(total["saldo_awal"]),
        _num(total["saldo_akhir"]), _num(total["selisih"]),
    ])
    for c in d[d.max_row]:
        c.font = Font(bold=True)
    return d


def _tgl_sel(v):
    """Tanggal untuk sel Excel: teks dd/mm/YYYY (konvensi sheet lain), None→kosong."""
    return v.strftime("%d/%m/%Y") if v else ""


def bonus_sheet(wb, data, title, rentang_label=""):
    """Sheet Rekonsiliasi Bonus — SATU tabel untuk semua ember.

    `data` = keluaran `web.bonus.rekonsiliasi_bonus`. Halaman /bonus/ punya tiga
    tab karena itu navigasi layar; di Excel tiga sheet justru menyulitkan —
    klien bekerja dengan filter kolom. Jadi satu tabel, dan kolom `Status`
    (`Cocok`/`Hanya Panel`/`Hanya Bracket`/`Agregat`/`Agregat (selisih)`) yang
    membedakan ember.

    Baris **agregat** (lump bracket per kategori per hari) muat di kolom yang
    sama tanpa dipaksakan: `Username` kosong karena memang tak ada nama pemain,
    `Kategori` = kategori BRACKET, dan kedua kolom nominal = total masing-masing
    sisi. Tanggalnya adalah tanggal EFEKTIF (hasil "TGL dd.mm.yyyy" bila ada),
    sama dengan yang tampil di halaman.

    Konvensi tanda seragam: `Selisih` = Nominal Panel − Nominal Bracket, persis
    seperti `selisih` pada entri agregat. Hanya-panel positif, hanya-bracket
    negatif. Baris `Cocok` boleh berselisih bukan nol — pairing berkunci
    `int(abs(nominal))`, jadi sen kedua sisi bisa berbeda; angkanya dibiarkan
    apa adanya, bukan dinolkan.

    Kunci `agregat` dibaca lewat `.get` karena untuk toko Nexus kunci itu SAMA
    SEKALI tak ada di `data` (lihat `web/bonus.py`).
    """
    d = wb.create_sheet(title)
    if rentang_label:
        d.append([rentang_label])
        d["A1"].font = Font(bold=True)
    d.append(["Tanggal", "Username", "Kategori", "Nominal Panel",
              "Nominal Bracket", "Selisih", "Status"])
    for c in d[d.max_row]:
        c.font = Font(bold=True)

    nol = Decimal("0")
    tot_panel = tot_bracket = nol
    for c in data["cocok"]:
        p, b = c["panel"], c["bracket"]
        panel, bracket = p["nominal"] or nol, b["nominal"] or nol
        tot_panel, tot_bracket = tot_panel + panel, tot_bracket + bracket
        d.append([_tgl_sel(p["tanggal"]), p["username"], p["kategori_detail"],
                  _num(panel), _num(bracket), _num(panel - bracket), "Cocok"])
    for r in data["panel_only"]:
        panel = r["nominal"] or nol
        tot_panel += panel
        d.append([_tgl_sel(r["tanggal"]), r["username"], r["kategori_detail"],
                  _num(panel), _num(None), _num(panel), "Hanya Panel"])
    for r in data["bracket_only"]:
        bracket = r["nominal"] or nol
        tot_bracket += bracket
        d.append([_tgl_sel(r["tanggal"]), r["username"], r["kategori_detail"],
                  _num(None), _num(bracket), _num(-bracket), "Hanya Bracket"])
    for e in data.get("agregat") or ():
        tot_panel += e["panel_total"]
        tot_bracket += e["bracket_total"]
        d.append([_tgl_sel(e["tanggal"]), "", e["kategori"],
                  _num(e["panel_total"]), _num(e["bracket_total"]),
                  _num(e["selisih"]),
                  "Agregat" if e["status"] == "cocok" else "Agregat (selisih)"])

    d.append(["TOTAL", "", "", _num(tot_panel), _num(tot_bracket),
              _num(tot_panel - tot_bracket), ""])
    for c in d[d.max_row]:
        c.font = Font(bold=True)
    return d


def _sheet_title(base, existing):
    """Judul sheet <=31 char, tanpa karakter terlarang openpyxl, anti-duplikat."""
    t = re.sub(r"[\\/*?:\[\]]", "-", base)[:31]
    n, out = 2, t
    while out in existing:
        suffix = f" ({n})"
        out = t[: 31 - len(suffix)] + suffix
        n += 1
    existing.add(out)
    return out


def results_sheet(wb, run, title, rel_labels):
    """Tulis satu sheet 'Hasil' untuk `run` (kolom identik dgn export_run lama)."""
    L, R = rel_labels.get(run.relation, _REL_LABELS_FALLBACK)
    d = wb.create_sheet(title)
    headers = ["Status", f"{L} Ticket", f"{L} Nominal", f"{L} Username", f"{L} Nama Lengkap",
               f"{L} Player Bank", f"{L} Bank Title", f"{L} Handler", f"{L} Waktu",
               R, f"{R} Sumber", f"{R} Nominal", f"{R} Waktu", "Skor", "Alasan", "Detail"]
    d.append(headers)
    for c in d[1]:
        c.font = Font(bold=True)
    qs = run.results.select_related("left", "right", "left__source_type", "right__source_type")
    for r in qs.iterator():
        left, right = r.left, r.right
        d.append([
            r.get_bucket_display(),
            left.ticket_no if left else "",
            float(left.amount) if left else "",
            left.username if left else "",
            left.counterparty if left else "",
            (left.raw or {}).get("Player Bank", "") if left else "",
            (left.raw or {}).get("Bank Title", "") if left else "",
            (left.raw or {}).get("Handler", "") if left else "",
            left.occurred_at.strftime("%d/%m %H:%M") if left and left.occurred_at else "",
            (right.ticket_no or right.counterparty) if right else "",
            right.source_type.key if right else "",
            float(right.amount) if right else "",
            right.occurred_at.strftime("%d/%m %H:%M") if right and right.occurred_at else "",
            round(r.score or 0),
            reason_label(r.reason_code),
            r.reason_detail,
        ])
    return d


def build_batch_workbook(batch, batch_no, rel_labels):
    """Workbook satu batch: sheet Ringkasan + satu sheet Hasil per run."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    s = batch.summary or {}
    dp, wd, buckets = s.get("dp", {}), s.get("wd", {}), s.get("buckets", {})
    rows = [
        ("Toko", batch.toko.name if batch.toko else ""),
        ("Batch", f"#{batch_no}" if batch_no else f"#{batch.pk}"),
        ("Tanggal rekonsiliasi", batch.recon_date.strftime("%d/%m/%Y") if batch.recon_date else ""),
        ("Toleransi", f"{batch.tolerance.name} (±{batch.tolerance.date_window_days} hari)"),
        ("Dibuat", batch.created_at.strftime("%d/%m/%Y %H:%M")),
        # Stempel versi: bila hasil lama perlu ditelusuri ulang, ketahuan versi
        # aplikasi mana yang membuatnya (aturan pencocokan berbeda antar versi).
        ("Versi aplikasi", f"v{app_versi()}"),
        ("", ""),
        ("DP Panel", dp.get("panel", 0)),
        ("DP Uang (matched)", dp.get("money_matched", dp.get("money", 0))),
        ("DP Selisih", dp.get("selisih", 0)),
        ("WD Panel", wd.get("panel", 0)),
        ("WD Uang (matched)", wd.get("money_matched", wd.get("money", 0))),
        ("WD Selisih", wd.get("selisih", 0)),
        ("", ""),
        ("Cocok", buckets.get("cocok", 0)),
        ("Perlu Ditinjau", buckets.get("perlu_tinjau", 0)),
        ("Tidak Cocok", buckets.get("tidak_cocok", 0)),
    ]
    for label, val in rows:
        ws.append([label, val])
    for cell in ws["A"]:
        cell.font = Font(bold=True)

    from reconciliation.engine import filter_visible_runs

    titles = {"Ringkasan"}
    for run in filter_visible_runs(batch.runs.all().select_related("tolerance")):
        results_sheet(wb, run, _sheet_title(f"Hasil {run.get_relation_display()}", titles), rel_labels)

    # Sheet tambahan (query-time, retroaktif): Breakdown Bracket + Rincian
    # Rekening + Rekonsiliasi Bonus untuk TANGGAL batch. Hanya ditambahkan bila
    # ada baris — batch tanpa data FR/rekening/bonus tetap identik format lama
    # (Ringkasan + Hasil).
    if batch.recon_date and batch.toko_id:
        from web.bonus import rekonsiliasi_bonus
        from web.breakdown import bracket_breakdown
        from web.rekening import rekening_breakdown

        tgl_label = batch.recon_date.strftime("%d/%m/%Y")
        bd = bracket_breakdown(batch.toko, batch.recon_date)
        if bd["count"]:
            breakdown_sheet(wb, bd, _sheet_title("Breakdown Bracket", titles), tgl_label)
        rk = rekening_breakdown(batch.toko, batch.recon_date)
        if rk["count"]:
            rekening_sheet(wb, rk, _sheet_title("Rincian Rekening", titles))
        bn = rekonsiliasi_bonus(batch.toko, batch.recon_date, batch.recon_date)
        if (bn["cocok"] or bn["panel_only"] or bn["bracket_only"]
                or bn.get("agregat")):
            bonus_sheet(wb, bn, _sheet_title("Rekonsiliasi Bonus", titles), tgl_label)
    return wb

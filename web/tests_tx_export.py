"""Export Excel Transaksi mengikuti filter aktif + guard batas baris."""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction

User = get_user_model()
_seq = iter(range(1, 100000))


class TxExportTests(TestCase):
    def setUp(self):
        User.objects.create_user("aud", "a@a.co", "pw12345", role="supervisor")
        self.client.login(username="aud", password="pw12345")
        self.lbs = Toko.objects.get(key="lbs")
        self.bank = SourceType.objects.get_or_create(key="bank", defaults={"name": "Bank"})[0]
        self.up = Upload.objects.create(source_type=self.bank, toko=self.lbs)
        self.client.post(reverse("set_toko"), {"toko_id": self.lbs.id})

    def tx(self, amount, **kw):
        return Transaction.objects.create(
            upload=self.up, source_type=self.bank, toko=self.lbs, jenis="depo",
            amount=Decimal(str(amount)), money_delta=Decimal(str(amount)),
            occurred_at=datetime(2026, 6, 27, 10, 0), row_hash=f"e-{next(_seq)}", **kw,
        )

    def _wb(self, resp):
        return load_workbook(BytesIO(resp.content))

    def test_export_menghormati_filter_q(self):
        self.tx(10000, counterparty="ANDI")
        self.tx(20000, counterparty="BUDI")
        r = self.client.get(reverse("transactions"), {"export": "1", "q": "ANDI"})
        self.assertEqual(
            r["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        ws = self._wb(r).active
        body = list(ws.iter_rows(min_row=2, values_only=True))
        joined = "\n".join(str(row) for row in body)
        self.assertIn("ANDI", joined)
        self.assertNotIn("BUDI", joined)

    def test_header_kolom_benar(self):
        self.tx(10000, counterparty=" X")
        r = self.client.get(reverse("transactions"), {"export": "1"})
        ws = self._wb(r).active
        header = [c.value for c in ws[1]]
        self.assertEqual(
            header,
            ["Waktu", "Sumber", "Jenis", "Nominal", "Δ Uang", "Ticket",
             "Username", "Nama Lengkap", "Nama di Bank"],
        )

    def test_export_kosong_hanya_header(self):
        r = self.client.get(reverse("transactions"), {"export": "1", "q": "TIDAKADA"})
        ws = self._wb(r).active
        self.assertEqual(ws.max_row, 1)

    def test_guard_batas_baris(self):
        from web import views
        self.tx(10000, counterparty="A")
        self.tx(20000, counterparty="B")
        orig = views.TX_EXPORT_LIMIT
        views.TX_EXPORT_LIMIT = 1
        try:
            r = self.client.get(reverse("transactions"), {"export": "1"}, follow=True)
        finally:
            views.TX_EXPORT_LIMIT = orig
        self.assertContains(r, "persempit")

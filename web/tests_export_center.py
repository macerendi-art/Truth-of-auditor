"""Menu Export: per-tanggal / per-toko / semua toko (admin) — bulk = ZIP berisi
xlsx per-(toko,tanggal); nama file memuat toko + tanggal. Export bulanan =
ringkasan + Breakdown Bracket / Rincian Rekening / Bonus sebulan.
"""
import io
import zipfile
from datetime import date, datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from reconciliation.models import MatchResult, MatchRun, ReconBatch, ToleranceProfile
from sources.models import SourceType, Toko, Upload
from transactions.models import Transaction

User = get_user_model()
_seq = iter(range(1, 100000))

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ExportCenterBase(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            "adm", "adm@a.co", "pw12345", role="admin"
        )
        self.client.login(username="adm", password="pw12345")
        self.lbs = Toko.objects.get(key="lbs")
        self.slo = Toko.objects.get(key="slo")
        self.tol = ToleranceProfile.objects.get_or_create(
            name="Default", defaults={"date_window_days": 1}
        )[0]
        self.panel = SourceType.objects.get_or_create(key="panel", defaults={"name": "Panel"})[0]

    def _batch(self, toko, recon_date, with_result=True):
        batch = ReconBatch.objects.create(
            toko=toko, tolerance=self.tol, recon_date=recon_date,
            summary={"dp": {"panel": 100, "money_matched": 100, "selisih": 0},
                     "wd": {"panel": 0, "money_matched": 0, "selisih": 0},
                     "buckets": {"cocok": 1, "perlu_tinjau": 0, "tidak_cocok": 0}},
        )
        run = MatchRun.objects.create(
            relation=MatchRun.Relation.PANEL_BANK, tolerance=self.tol, batch=batch,
            summary={"left": 1, "cocok": 1, "perlu_tinjau": 0, "tidak_cocok": 0},
        )
        if with_result:
            up = Upload.objects.create(source_type=self.panel, toko=toko)
            left = Transaction.objects.create(
                upload=up, source_type=self.panel, toko=toko, jenis="depo",
                amount=Decimal("50000"), ticket_no=f"D-{next(_seq)}",
                row_hash=f"x-{next(_seq)}", raw={},
            )
            MatchResult.objects.create(
                run=run, bucket=MatchResult.Bucket.COCOK, left=left, reason_code="ticket_exact",
            )
        return batch


class ExportSingleTests(ExportCenterBase):
    def test_satu_batch_langsung_xlsx(self):
        self._batch(self.lbs, date(2026, 6, 27))
        r = self.client.get(reverse("export_center"),
                            {"toko": self.lbs.id, "from": "2026-06-27"})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], XLSX_CT)
        cd = r["Content-Disposition"]
        self.assertIn("2026-06-27", cd)
        self.assertIn(self.lbs.name.replace(" ", "_"), cd)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertIn("Ringkasan", wb.sheetnames)
        self.assertTrue(any(s.startswith("Hasil") for s in wb.sheetnames))

    def test_form_get_tanpa_param_render(self):
        r = self.client.get(reverse("export_center"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Export")

    def test_sampai_saja_dianggap_satu_tanggal(self):
        # isi "Sampai" saja -> diperlakukan tanggal itu juga -> 1 batch xlsx
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.lbs, date(2026, 6, 28))
        r = self.client.get(reverse("export_center"),
                            {"toko": self.lbs.id, "to": "2026-06-27"})
        self.assertEqual(r["Content-Type"], XLSX_CT)
        self.assertIn("2026-06-27", r["Content-Disposition"])

    def test_kosong_redirect_message(self):
        r = self.client.get(reverse("export_center"),
                            {"toko": self.lbs.id, "from": "2026-01-01"}, follow=True)
        self.assertContains(r, "Tidak ada batch")


class ExportBulkTests(ExportCenterBase):
    def test_rentang_dua_batch_jadi_zip_per_tanggal(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.lbs, date(2026, 6, 28))
        r = self.client.get(reverse("export_center"),
                            {"toko": self.lbs.id, "from": "2026-06-27", "to": "2026-06-28"})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/zip")
        zf = zipfile.ZipFile(io.BytesIO(r.content))
        names = sorted(zf.namelist())
        self.assertEqual(len(names), 2)
        safe = self.lbs.name.replace(" ", "_")
        self.assertIn(f"rekonsiliasi_{safe}_2026-06-27.xlsx", names)
        self.assertIn(f"rekonsiliasi_{safe}_2026-06-28.xlsx", names)
        # tiap member = workbook valid dgn sheet Ringkasan + Hasil
        wb = load_workbook(io.BytesIO(zf.read(names[0])))
        self.assertIn("Ringkasan", wb.sheetnames)
        self.assertTrue(any(s.startswith("Hasil") for s in wb.sheetnames))

    def test_semua_toko_admin(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.slo, date(2026, 6, 27))
        r = self.client.get(reverse("export_center"), {"toko": "all", "from": "2026-06-27"})
        self.assertEqual(r["Content-Type"], "application/zip")
        zf = zipfile.ZipFile(io.BytesIO(r.content))
        self.assertEqual(len(zf.namelist()), 2)

    def test_semua_tanggal_satu_toko(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.lbs, date(2026, 6, 28))
        self._batch(self.lbs, date(2026, 6, 29))
        r = self.client.get(reverse("export_center"), {"toko": self.lbs.id})
        zf = zipfile.ZipFile(io.BytesIO(r.content))
        self.assertEqual(len(zf.namelist()), 3)


class ExportRBACTests(ExportCenterBase):
    def _login_auditor(self, *tokos):
        u = User.objects.create_user("aud2", "aud2@a.co", "pw12345", role="auditor")
        u.allowed_tokos.set(tokos)
        self.client.logout()
        self.client.login(username="aud2", password="pw12345")

    def test_auditor_semua_toko_ditolak(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._login_auditor(self.lbs)
        r = self.client.get(reverse("export_center"),
                            {"toko": "all", "from": "2026-06-27"}, follow=True)
        self.assertNotEqual(r.get("Content-Type"), "application/zip")
        self.assertContains(r, "admin")

    def test_auditor_toko_di_luar_allowed_ditolak(self):
        self._batch(self.slo, date(2026, 6, 27))
        self._login_auditor(self.lbs)
        r = self.client.get(reverse("export_center"),
                            {"toko": self.slo.id, "from": "2026-06-27"}, follow=True)
        self.assertNotEqual(r.get("Content-Type"), XLSX_CT)

    def test_auditor_toko_sendiri_boleh(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._login_auditor(self.lbs)
        r = self.client.get(reverse("export_center"),
                            {"toko": self.lbs.id, "from": "2026-06-27"})
        self.assertEqual(r["Content-Type"], XLSX_CT)

    def test_opsi_semua_toko_hanya_admin_di_form(self):
        r = self.client.get(reverse("export_center"))
        self.assertContains(r, "Semua toko")
        self._login_auditor(self.lbs)
        r = self.client.get(reverse("export_center"))
        self.assertNotContains(r, "Semua toko")


class ExportBulananTests(ExportCenterBase):
    """mode=bulanan → xlsx ringkasan 1 bulan (bukan detail batch)."""

    def test_form_tampil_section_bulanan(self):
        r = self.client.get(reverse("export_center"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Export Bulanan")
        self.assertContains(r, 'name="mode"')
        self.assertContains(r, 'value="bulanan"')
        self.assertContains(r, 'name="bulan"')

    def test_satu_toko_xlsx_summary(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.lbs, date(2026, 6, 28))
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-06",
        })
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], XLSX_CT)
        cd = r["Content-Disposition"]
        self.assertIn("ringkasan_bulanan_", cd)
        self.assertIn("2026-06", cd)
        self.assertIn(self.lbs.name.replace(" ", "_"), cd)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertIn("Ringkasan Bulanan", wb.sheetnames)
        ws = wb["Ringkasan Bulanan"]
        # meta + header + 2 baris harian + TOTAL
        cells_a = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
        self.assertIn("Tanggal", cells_a)
        self.assertIn("TOTAL", cells_a)
        # baris data 27 & 28 Juni
        tgls = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
        self.assertIn("27/06/2026", tgls)
        self.assertIn("28/06/2026", tgls)
        # TOTAL cocok = 1+1 dari fixture _batch
        total_row = ws.max_row
        self.assertEqual(ws.cell(row=total_row, column=1).value, "TOTAL")
        self.assertEqual(ws.cell(row=total_row, column=8).value, 2)  # Cocok

    def test_bulan_kosong_redirect(self):
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-01",
        }, follow=True)
        self.assertContains(r, "Tidak ada ringkasan bulanan")

    def test_bulan_invalid_redirect(self):
        self._batch(self.lbs, date(2026, 6, 27))
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-13",
        }, follow=True)
        self.assertContains(r, "Bulan tidak valid")

    def test_semua_toko_admin_zip(self):
        self._batch(self.lbs, date(2026, 6, 27))
        self._batch(self.slo, date(2026, 6, 15))
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": "all", "bulan": "2026-06",
        })
        self.assertEqual(r["Content-Type"], "application/zip")
        zf = zipfile.ZipFile(io.BytesIO(r.content))
        names = sorted(zf.namelist())
        self.assertEqual(len(names), 2)
        self.assertTrue(all(n.startswith("ringkasan_bulanan_") for n in names))
        self.assertTrue(all("2026-06" in n for n in names))
        wb = load_workbook(io.BytesIO(zf.read(names[0])))
        self.assertIn("Ringkasan Bulanan", wb.sheetnames)

    def test_auditor_semua_toko_ditolak(self):
        self._batch(self.lbs, date(2026, 6, 27))
        u = User.objects.create_user("aud_m", "audm@a.co", "pw12345", role="auditor")
        u.allowed_tokos.set([self.lbs])
        self.client.logout()
        self.client.login(username="aud_m", password="pw12345")
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": "all", "bulan": "2026-06",
        }, follow=True)
        self.assertNotEqual(r.get("Content-Type"), "application/zip")
        self.assertContains(r, "admin")

    def test_audit_log_export_bulanan(self):
        from core.models import AuditLog
        self._batch(self.lbs, date(2026, 6, 27))
        self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-06",
        })
        self.assertTrue(AuditLog.objects.filter(aksi="export_bulanan").exists())

    def test_xlsx_memuat_breakdown_rekening_bonus(self):
        """Export bulanan + FR/mutasi/bonus sebulan → 4 sheet (sama builder harian)."""
        self._batch(self.lbs, date(2026, 6, 15))
        self._batch(self.lbs, date(2026, 6, 28))

        bracket = SourceType.objects.get_or_create(
            key="bracket", defaults={"name": "Bracket"}
        )[0]
        bank = SourceType.objects.get_or_create(key="bank", defaults={"name": "Bank"})[0]
        panel_bonus = SourceType.objects.get(key="panel_bonus")
        bracket_bonus = SourceType.objects.get(key="bracket_bonus")

        # FR Control Bracket — 2 hari di Juni (kategori DP & Bonus)
        up_br = Upload.objects.create(source_type=bracket, toko=self.lbs)
        for tgl, kat, total, saldo in (
            (date(2026, 6, 10), "Deposit", "100000", "500000"),
            (date(2026, 6, 20), "Bonus", "25000", "525000"),
        ):
            n = next(_seq)
            Transaction.objects.create(
                upload=up_br, source_type=bracket, toko=self.lbs, jenis="lainnya",
                amount=abs(Decimal(total)), money_delta=Decimal(total),
                balance_after=Decimal(saldo), posted_date=tgl,
                occurred_at=datetime(tgl.year, tgl.month, tgl.day, 10, 0),
                row_hash=f"fr-m-{n}",
                raw={"Bank": "QRIS HOKI", "Kategori": kat, "Jam": "10:00"},
            )

        # Mutasi bank
        up_bk = Upload.objects.create(
            source_type=bank, toko=self.lbs, provider="BCA", owner_name="CM TEST"
        )
        tgl_m = date(2026, 6, 12)
        Transaction.objects.create(
            upload=up_bk, source_type=bank, toko=self.lbs, jenis="depo",
            amount=Decimal("75000"), money_delta=Decimal("75000"),
            balance_after=Decimal("1000000"),
            occurred_at=datetime(tgl_m.year, tgl_m.month, tgl_m.day, 11, 0),
            row_hash=f"bk-m-{next(_seq)}", raw={},
        )

        # Bonus panel + bracket (kategori Lucky Draw) di bulan yang sama
        up_pb = Upload.objects.create(source_type=panel_bonus, toko=self.lbs)
        up_bb = Upload.objects.create(source_type=bracket_bonus, toko=self.lbs)
        tgl_b = date(2026, 6, 18)
        for up, st, user in (
            (up_pb, panel_bonus, "Cici"),
            (up_bb, bracket_bonus, "cici"),
        ):
            n = next(_seq)
            Transaction.objects.create(
                upload=up, source_type=st, toko=self.lbs, jenis="bonus",
                amount=Decimal("40000"), money_delta=Decimal("0"), ticket_no="",
                username=user, posted_date=tgl_b,
                occurred_at=datetime(tgl_b.year, tgl_b.month, tgl_b.day, 10, 0),
                description=f"Lucky Draw {user}",
                raw={"Kategori": "Lucky Draw"}, row_hash=f"bn-m-{n}",
            )

        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-06",
        })
        self.assertEqual(r["Content-Type"], XLSX_CT)
        wb = load_workbook(io.BytesIO(r.content))
        names = wb.sheetnames
        self.assertIn("Ringkasan Bulanan", names)
        self.assertIn("Breakdown Bracket", names)
        self.assertIn("Rincian Rekening", names)
        self.assertIn("Rekonsiliasi Bonus", names)

        # Breakdown: caption rentang sebulan + FR Account
        bd = wb["Breakdown Bracket"]
        self.assertIn("01/06/2026", str(bd["A1"].value))
        self.assertIn("30/06/2026", str(bd["A1"].value))
        # cari nama akun di kolom B
        found_fr = any(
            row[1] and "QRIS HOKI" in str(row[1])
            for row in bd.iter_rows(values_only=True)
            if row and len(row) > 1
        )
        self.assertTrue(found_fr)

        # Rekening: label bank
        rk = wb["Rincian Rekening"]
        found_rk = any(
            row[1] and "BCA" in str(row[1]) and "CM TEST" in str(row[1])
            for row in rk.iter_rows(values_only=True)
            if row and len(row) > 1
        )
        self.assertTrue(found_rk)

        # Bonus: kolom Kategori + status Cocok
        bn = wb["Rekonsiliasi Bonus"]
        headers = None
        for row in bn.iter_rows(values_only=True):
            if row and row[0] == "Tanggal":
                headers = list(row)
                break
        self.assertEqual(
            headers,
            ["Tanggal", "Username", "Kategori", "Nominal Panel",
             "Nominal Bracket", "Selisih", "Status"],
        )
        statuses = [
            row[6] for row in bn.iter_rows(values_only=True)
            if row and row[6] in ("Cocok", "Hanya Panel", "Hanya Bracket")
        ]
        self.assertIn("Cocok", statuses)
        kats = [
            row[2] for row in bn.iter_rows(values_only=True)
            if row and row[2] and row[0] not in (None, "Tanggal", "TOTAL")
            and not str(row[0]).startswith("01/")
        ]
        # baris data punya kategori Lucky Draw
        self.assertTrue(any("Lucky" in str(k) for k in kats))

    def test_tanpa_fr_rekening_bonus_hanya_ringkasan(self):
        """Tanpa data sumber tambahan → workbook tetap 1 sheet (tak error)."""
        self._batch(self.lbs, date(2026, 6, 5))
        r = self.client.get(reverse("export_center"), {
            "mode": "bulanan", "toko": self.lbs.id, "bulan": "2026-06",
        })
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.sheetnames, ["Ringkasan Bulanan"])
